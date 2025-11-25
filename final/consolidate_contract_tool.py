"""
LANGGRAPH TOOL: Consolidate Contract
Consolidate duplicate rules in a data contract to create a master/golden contract

This tool takes a messy consumer contract with potential duplicates and creates
a clean master contract with unique, consolidated rules.
"""

import boto3
import json
import pandas as pd
import os
from typing import Dict, List, Any, Optional
from datetime import datetime
from pydantic import BaseModel, Field
from langchain_core.tools import tool


# =============================================================================
# CONFIGURATION
# =============================================================================

MAX_RETRIES = 3

# Column name variations we should recognize (case-insensitive matching)
COLUMN_ALIASES = {
    'business_rule': ['business_rule', 'business rule', 'businessrule', 'rule', 'business_rules'],
    'business_term': ['business_term', 'business term', 'businessterm', 'term', 'business_terms'],
    'description': ['description', 'desc', 'descriptions']
}


# =============================================================================
# PYDANTIC MODELS FOR TOOL INPUT/OUTPUT
# =============================================================================

class FileInfo(BaseModel):
    """Information about a generated file"""
    path: str = Field(description="Full path to the file")
    filename: str = Field(description="Just the filename with extension")
    type: str = Field(description="Type of file: master_contract, audit_trail, or audit_trail_formatted")
    purpose: str = Field(description="Whether this is PRIMARY OUTPUT or AUDIT ONLY")
    row_count: int = Field(description="Number of rows in the file")
    preview: str = Field(description="Text preview of first 10 rows")


class ConsolidationSummary(BaseModel):
    """Statistics about the consolidation results"""
    total_original_rules: int = Field(description="Number of rules in input file")
    total_master_rules: int = Field(description="Number of unique rules in master contract")
    rules_consolidated: int = Field(description="Number of duplicate rules removed")
    low_confidence_consolidations: int = Field(description="Number of consolidations flagged for review (confidence < 0.9)")
    business_terms_processed: int = Field(description="Number of unique business terms processed")
    multi_rule_terms_processed: int = Field(description="Business terms with multiple rules that needed LLM analysis")
    single_rule_terms_processed: int = Field(description="Business terms with single rules (no LLM needed)")


class ConsolidateContractOutput(BaseModel):
    """Output from the consolidate_contract tool"""
    status: str = Field(description="'success' or 'error'")
    message: str = Field(description="Human-readable summary of what happened")
    files_created: Optional[List[FileInfo]] = Field(
        default=None,
        description="List of files created. First file is always master_contract (PRIMARY OUTPUT). "
                    "Subsequent files are audit trails (AUDIT ONLY - do not use as input to other functions)."
    )
    summary: Optional[ConsolidationSummary] = Field(
        default=None,
        description="Detailed statistics about consolidation results"
    )
    warnings: List[str] = Field(
        default_factory=list,
        description="Non-fatal warnings encountered during processing"
    )
    errors: List[str] = Field(
        default_factory=list,
        description="Partial errors (processing continued despite these)"
    )
    error_type: Optional[str] = Field(
        default=None,
        description="Type of error if status is 'error'"
    )
    suggestion: Optional[str] = Field(
        default=None,
        description="Suggestion for resolving the error"
    )


# =============================================================================
# UTILITY: FLEXIBLE COLUMN MATCHING
# =============================================================================

def find_column_name(df: pd.DataFrame, target_column: str) -> Optional[str]:
    """
    Find the actual column name in DataFrame that matches our target
    Handles case-insensitive matching and common variations
    
    Parameters:
        df: DataFrame to search
        target_column: Standard column name we're looking for
        
    Returns:
        Actual column name in df, or None if not found
    """
    if target_column not in COLUMN_ALIASES:
        # Not a column we have aliases for, try exact match
        return target_column if target_column in df.columns else None
    
    # Get list of possible names for this column
    possible_names = COLUMN_ALIASES[target_column]
    
    # Create lowercase version of all df columns for matching
    df_columns_lower = {col.lower().strip(): col for col in df.columns}
    
    # Try each possible name
    for possible_name in possible_names:
        possible_lower = possible_name.lower().strip()
        if possible_lower in df_columns_lower:
            return df_columns_lower[possible_lower]
    
    return None


def standardize_dataframe_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Rename DataFrame columns to standard names
    
    Parameters:
        df: DataFrame with potentially non-standard column names
        
    Returns:
        DataFrame with standardized column names
    """
    rename_map = {}
    
    # Find and map each standard column
    for standard_name in ['business_rule', 'business_term', 'description']:
        actual_name = find_column_name(df, standard_name)
        if actual_name and actual_name != standard_name:
            rename_map[actual_name] = standard_name
    
    if rename_map:
        df = df.rename(columns=rename_map)
    
    return df


# =============================================================================
# SYSTEM PROMPT
# =============================================================================

def get_system_prompt():
    """Returns the system prompt for rule consolidation"""
    return """You are a data quality expert analyzing business rules to identify and eliminate true duplicates. Your goal is to ensure each unique business logic constraint is preserved while removing redundant rules that say the same thing in different words.

## INPUT FORMAT

Business rules are formatted as "[Business Term]: [Rule Description]"

**CRITICAL**: If a rule starts with text that matches or is similar to the business term, ignore that prefix. Only analyze the actual rule logic after the colon.

Examples:
- "Account Category: The value must be chosen from an approved list" → Analyze: "The value must be chosen from an approved list"
- "FI - Muni/State Invested: This field is required" → Analyze: "This field is required"

## WHAT MAKES RULES DUPLICATES

Two rules are duplicates ONLY if they enforce the exact same business logic constraint. They must be identical in:
1. **What** is being constrained
2. **How** it is constrained  
3. **When** it applies (conditions/scope)

### DUPLICATES - Same constraint, different wording:
- "Field is required" ≡ "Field cannot be empty" ≡ "Field must not be null" ≡ "This field is required and cannot be empty"
- "Value must be from approved list" ≡ "Value must be chosen from a standard list" ≡ "The value must be chosen from an approved list"
- "CUSIP must be 9 characters" ≡ "CUSIP length must equal 9"

### NOT DUPLICATES - Different business logic:
- "Field is required" ≠ "Field is required only if X is Y" (unconditional vs conditional)
- "Field is required if X is Y" ≠ "Field is required if X is Z" (different conditions)
- "Value from approved list" ≠ "Field cannot be empty" (different constraints)
- "Value from list if OMS is AIM" ≠ "Value from list if OMS is BBGAIM" (different condition values)
- "Field > 0" ≠ "Field >= 0" (different boundaries)

## CONDITIONAL LOGIC IS CRITICAL

Rules with conditional qualifiers ("if", "when", "only if", "only when") are fundamentally different from unconditional rules.

**IMPORTANT**: A conditional rule is NEVER a duplicate of an unconditional rule, even if the constraint text is similar.

Examples of UNIQUE rules (not duplicates):
- "This field is required"
- "This field is required only if the Muni Inv Focus is state"

These define different business logic - one always applies, one conditionally applies.

## COMPLEMENTARY RULES

Multiple rules for the same business term may represent different aspects of validation. These are NOT duplicates:

Example - All three are UNIQUE:
- "If the OMS designation is AIM, this field cannot be null"
- "The value must be chosen from an approved list"
- "This field is required and cannot be empty"

These rules complement each other to form a complete validation contract. Keep them all unless they're truly redundant.

## CANONICAL RULE SELECTION

When rules ARE duplicates, select the clearest version as canonical:

Priority:
1. **Clarity**: Most unambiguous and explicit
2. **Completeness**: Captures full constraint without being verbose
3. **Standard phrasing**: Prefer "must" over "should", positive statements over negative

Example: For duplicates "required", "cannot be null", "must not be empty"
Select: "This field is required and cannot be empty" (most complete and clear)

## CONFIDENCE SCORING

- **1.0**: Absolutely certain (mathematically/logically identical)
- **0.9-0.99**: Very confident (clear semantic equivalence, minor wording differences)
- **0.8-0.89**: Confident (likely equivalent with slight ambiguity)
- **< 0.8**: When uncertain, mark as UNIQUE rather than duplicate

Be conservative. If there's any doubt about whether rules express the same business logic, keep them as unique.

## OUTPUT FORMAT

Return ONLY valid JSON with no markdown formatting or additional text:

{
    "business_term": "string",
    "total_rules_analyzed": 0,
    "unique_rules_identified": 0,
    "duplicates_found": false,
    "analysis": [
        {
            "original_rule": "exact input text",
            "original_rule_index": 0,
            "status": "unique|duplicate",
            "canonical_rule": "best version of rule",
            "confidence": 0.95,
            "consolidated_with": ["array of other rule texts if duplicate"],
            "duplicate_of_index": null,
            "reasoning": "clear explanation of decision"
        }
    ],
    "notes": "any additional observations"
}

Return analysis array in the same order as input rules."""


# =============================================================================
# DATA LOADING AND VALIDATION
# =============================================================================

def load_and_validate_contract(csv_path: str) -> tuple[pd.DataFrame, List[str]]:
    """
    Load and validate the input CSV file
    
    Parameters:
        csv_path: Path to input CSV file
        
    Returns:
        Tuple of (cleaned_df, warnings_list)
    """
    warnings = []
    
    # Check file exists
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"Input file not found: {csv_path}")
    
    # Load CSV
    df = pd.read_csv(csv_path)
    original_count = len(df)
    
    # Standardize column names
    df = standardize_dataframe_columns(df)
    
    # Check required columns exist after standardization
    required_columns = ['business_rule', 'business_term']
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"Required columns not found (even after flexible matching): {missing_columns}. "
                        f"Available columns: {list(df.columns)}")
    
    # Remove rows with null business_rule or business_term
    df = df.dropna(subset=['business_rule', 'business_term'])
    null_removed = original_count - len(df)
    if null_removed > 0:
        warnings.append(f"Removed {null_removed} rows with null values")
    
    # Trim whitespace
    df['business_rule'] = df['business_rule'].str.strip()
    df['business_term'] = df['business_term'].str.strip()
    
    # Remove exact duplicates
    before_dedup = len(df)
    df = df.drop_duplicates(subset=['business_rule', 'business_term'])
    duplicates_removed = before_dedup - len(df)
    if duplicates_removed > 0:
        warnings.append(f"Removed {duplicates_removed} exact duplicate rows")
    
    if len(df) == 0:
        raise ValueError("No valid data remaining after cleaning")
    
    return df, warnings


# =============================================================================
# BUSINESS TERM EXTRACTION
# =============================================================================

def extract_business_terms(df: pd.DataFrame) -> Dict[str, List[str]]:
    """
    Categorize business terms by rule count
    
    Parameters:
        df: Cleaned DataFrame
        
    Returns:
        Dict with 'multi_rule_terms' and 'single_rule_terms'
    """
    term_counts = df.groupby('business_term').size()
    
    multi_rule_terms = term_counts[term_counts >= 2].index.tolist()
    multi_rule_terms.sort()
    
    single_rule_terms = term_counts[term_counts == 1].index.tolist()
    single_rule_terms.sort()
    
    return {
        'multi_rule_terms': multi_rule_terms,
        'single_rule_terms': single_rule_terms
    }


def get_rules_for_business_term(df: pd.DataFrame, business_term: str) -> pd.DataFrame:
    """Filter dataframe to only rules for a specific business term"""
    return df[df['business_term'] == business_term].copy()


# =============================================================================
# LLM INVOCATION
# =============================================================================

def invoke_bedrock_model(system_prompt: str, user_prompt: str) -> str:
    """
    Call AWS Bedrock API to invoke Claude
    
    Parameters:
        system_prompt: System instructions
        user_prompt: User message
        
    Returns:
        Raw response text from model
    """
    bedrock = boto3.client(
        service_name='bedrock-runtime',
        region_name='us-east-1'
    )
    
    body = {
        "modelId": "global.anthropic.claude-haiku-4-5-20251001-v1:0",
        "inferenceConfig": {
            "maxTokens": 4096,
            "temperature": 0
        },
        "system": [{"text": system_prompt}],
        "messages": [{
            'role': 'user',
            'content': [{'text': user_prompt}]
        }]
    }
    
    response = bedrock.converse(**body)
    result = response['output']['message']['content'][0]['text']
    
    return result


def validate_llm_response_structure(response: Dict, business_term: str):
    """Validate that LLM response has expected structure"""
    if 'business_term' not in response:
        raise ValueError("Missing 'business_term' field")
    
    if 'analysis' not in response:
        raise ValueError("Missing 'analysis' field")
    
    if not isinstance(response['analysis'], list):
        raise ValueError("'analysis' must be a list")
    
    for i, entry in enumerate(response['analysis']):
        required_fields = ['original_rule', 'status', 'canonical_rule', 'confidence']
        for field in required_fields:
            if field not in entry:
                raise ValueError(f"Analysis entry {i} missing field: {field}")
        
        if entry['status'] not in ['unique', 'duplicate']:
            raise ValueError(f"Analysis entry {i} has invalid status: {entry['status']}")
        
        if not (0.0 <= entry['confidence'] <= 1.0):
            raise ValueError(f"Analysis entry {i} has invalid confidence: {entry['confidence']}")


# =============================================================================
# LLM CONSOLIDATION
# =============================================================================

def consolidate_rules_with_llm(rules_df: pd.DataFrame, business_term: str) -> Dict[str, Any]:
    """
    Use LLM to semantically consolidate duplicate rules for a business term
    
    Parameters:
        rules_df: All rules for one business term
        business_term: The business term being processed
        
    Returns:
        Dictionary with consolidation analysis
    """
    rules_list = rules_df['business_rule'].tolist()
    
    user_prompt = f"""Business Term: {business_term}

Rules to analyze:
{json.dumps(rules_list, indent=2)}

Analyze these rules and identify semantic duplicates. Return your analysis as valid JSON only (no markdown, no code blocks)."""
    
    system_prompt = get_system_prompt()
    
    for attempt in range(MAX_RETRIES):
        try:
            llm_response = invoke_bedrock_model(system_prompt, user_prompt)
            
            # Clean markdown if present
            llm_response = llm_response.strip()
            if llm_response.startswith("```"):
                lines = llm_response.split('\n')
                llm_response = '\n'.join(lines[1:-1])
            
            result = json.loads(llm_response)
            validate_llm_response_structure(result, business_term)
            
            return result
            
        except (json.JSONDecodeError, ValueError, KeyError) as e:
            if attempt < MAX_RETRIES - 1:
                user_prompt += "\n\nIMPORTANT: Return ONLY valid JSON. No markdown code blocks, no explanations outside JSON."
            else:
                # Return error structure
                return {
                    "business_term": business_term,
                    "error": f"Failed to parse LLM response after {MAX_RETRIES} attempts: {str(e)}",
                    "total_rules_analyzed": len(rules_list),
                    "unique_rules_identified": 0,
                    "duplicates_found": False,
                    "analysis": []
                }


def create_single_rule_audit_entry(business_term: str, rule_text: str) -> Dict[str, Any]:
    """Create audit entry for business terms with only one rule (no LLM needed)"""
    return {
        "business_term": business_term,
        "total_rules_analyzed": 1,
        "unique_rules_identified": 1,
        "duplicates_found": False,
        "analysis": [
            {
                "original_rule": rule_text,
                "original_rule_index": 0,
                "status": "unique",
                "canonical_rule": rule_text,
                "confidence": 1.0,
                "consolidated_with": [],
                "duplicate_of_index": None,
                "reasoning": "Only one rule detected for this business term. No consolidation needed."
            }
        ],
        "notes": "Single rule - no LLM processing required"
    }


# =============================================================================
# ORCHESTRATION
# =============================================================================

def process_all_business_terms(df: pd.DataFrame, 
                                term_categories: Dict[str, List[str]]) -> List[Dict[str, Any]]:
    """
    Process all business terms - both multi-rule (with LLM) and single-rule (without LLM)
    
    Parameters:
        df: Cleaned full dataset
        term_categories: Dict with 'multi_rule_terms' and 'single_rule_terms'
        
    Returns:
        List of consolidation results for all terms
    """
    all_results = []
    
    # Process multi-rule terms (needs LLM)
    multi_rule_terms = term_categories['multi_rule_terms']
    for business_term in multi_rule_terms:
        term_rules_df = get_rules_for_business_term(df, business_term)
        result = consolidate_rules_with_llm(term_rules_df, business_term)
        all_results.append(result)
    
    # Process single-rule terms (no LLM needed)
    single_rule_terms = term_categories['single_rule_terms']
    for business_term in single_rule_terms:
        term_rules_df = get_rules_for_business_term(df, business_term)
        if len(term_rules_df) == 1:
            rule_text = term_rules_df.iloc[0]['business_rule']
            result = create_single_rule_audit_entry(business_term, rule_text)
            all_results.append(result)
    
    return all_results


# =============================================================================
# MASTER CONTRACT CREATION
# =============================================================================

def create_master_contract(all_results: List[Dict[str, Any]]) -> tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    """
    Create master contract CSV and detailed audit CSV from LLM results
    
    Parameters:
        all_results: Consolidation results from all business terms
        
    Returns:
        Tuple of (master_df, audit_df, errors_list)
    """
    master_rules = []
    audit_records = []
    errors = []
    
    for result in all_results:
        business_term = result['business_term']
        
        # Track errors but continue processing
        if 'error' in result:
            errors.append(f"{business_term}: {result['error']}")
            continue
        
        seen_canonical = set()
        
        for rule_analysis in result['analysis']:
            # Create audit record
            audit_record = {
                'original_rule': rule_analysis['original_rule'],
                'business_term': business_term,
                'status': rule_analysis['status'],
                'canonical_rule': rule_analysis['canonical_rule'],
                'confidence': rule_analysis['confidence'],
                'needs_review': rule_analysis['confidence'] < 0.9,
                'reasoning': rule_analysis['reasoning']
            }
            audit_records.append(audit_record)
            
            # Add unique rules to master
            canonical_rule = rule_analysis['canonical_rule']
            if canonical_rule not in seen_canonical:
                seen_canonical.add(canonical_rule)
                
                original_rule_count = sum(
                    1 for r in result['analysis'] 
                    if r['canonical_rule'] == canonical_rule
                )
                
                consolidation_method = 'single_rule' if len(result['analysis']) == 1 else 'llm_consolidated'
                
                master_rules.append({
                    'business_rule': canonical_rule,
                    'business_term': business_term,
                    'consolidation_method': consolidation_method,
                    'original_rule_count': original_rule_count
                })
    
    master_df = pd.DataFrame(master_rules)
    audit_df = pd.DataFrame(audit_records)
    
    if not master_df.empty:
        master_df = master_df.sort_values(['business_term', 'business_rule']).reset_index(drop=True)
    
    if not audit_df.empty:
        audit_df = audit_df.sort_values(
            ['business_term', 'needs_review', 'original_rule'],
            ascending=[True, False, True]
        ).reset_index(drop=True)
    
    return master_df, audit_df, errors


# =============================================================================
# FILE PREVIEW
# =============================================================================

def generate_file_preview(df: pd.DataFrame, max_rows: int = 10) -> str:
    """Generate a text preview of the first N rows of a dataframe"""
    preview_df = df.head(max_rows)
    return preview_df.to_string(index=False)


# =============================================================================
# OUTPUT GENERATION
# =============================================================================

def save_outputs(master_df: pd.DataFrame, 
                audit_df: pd.DataFrame, 
                output_dir: str,
                timestamp: str) -> Dict[str, Any]:
    """
    Save master contract and audit trail
    
    Parameters:
        master_df: Consolidated master contract
        audit_df: Detailed audit trail
        output_dir: Directory to save outputs
        timestamp: Timestamp string for unique filenames
        
    Returns:
        Dictionary with file information and previews
    """
    os.makedirs(output_dir, exist_ok=True)
    
    files_created = []
    
    # Save master contract (PRIMARY OUTPUT)
    master_filename = f'master_contract_{timestamp}.csv'
    master_path = os.path.join(output_dir, master_filename)
    master_df.to_csv(master_path, index=False)
    
    files_created.append({
        'path': master_path,
        'filename': master_filename,
        'type': 'master_contract',
        'purpose': 'PRIMARY OUTPUT - Use this as input to other functions',
        'row_count': len(master_df),
        'preview': generate_file_preview(master_df, max_rows=10)
    })
    
    # Save audit trail (AUDIT ONLY)
    audit_filename = f'consolidation_audit_{timestamp}.csv'
    audit_path = os.path.join(output_dir, audit_filename)
    audit_df.to_csv(audit_path, index=False)
    
    files_created.append({
        'path': audit_path,
        'filename': audit_filename,
        'type': 'audit_trail',
        'purpose': 'AUDIT ONLY - Do NOT use as input to other functions. For human review only.',
        'row_count': len(audit_df),
        'preview': generate_file_preview(audit_df, max_rows=10)
    })
    
    # Try to save formatted Excel (AUDIT ONLY)
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        
        excel_filename = f'consolidation_audit_formatted_{timestamp}.xlsx'
        excel_path = os.path.join(output_dir, excel_filename)
        audit_df.to_excel(excel_path, index=False, engine='openpyxl')
        
        # Apply formatting
        wb = load_workbook(excel_path)
        ws = wb.active
        
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        needs_review_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'needs_review':
                needs_review_col = idx
                break
        
        if needs_review_col:
            for row in range(2, ws.max_row + 1):
                if ws.cell(row, needs_review_col).value == True:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row, col).fill = yellow_fill
        
        wb.save(excel_path)
        
        files_created.append({
            'path': excel_path,
            'filename': excel_filename,
            'type': 'audit_trail_formatted',
            'purpose': 'AUDIT ONLY - Formatted Excel with highlighting. Do NOT use as input to other functions.',
            'row_count': len(audit_df),
            'preview': 'Excel file with yellow highlighting for low-confidence consolidations'
        })
        
    except ImportError:
        pass  # openpyxl not available, skip Excel
    
    return files_created


# =============================================================================
# MAIN TOOL FUNCTION - LANGGRAPH @tool DECORATOR
# =============================================================================

@tool
def consolidate_contract(
    input_csv_path: str,
    output_dir: str = "./output",
    generate_excel: bool = True
) -> dict:
    """
    Consolidate duplicate rules in a data contract to create a master/golden contract.
    
    Takes a messy consumer contract that may contain duplicate or semantically identical rules 
    and uses LLM analysis to identify truly unique business logic. Outputs a clean master contract
    with consolidated rules plus detailed audit trails.
    
    IMPORTANT: 
    - The master_contract file is the PRIMARY OUTPUT - use this as input to other functions
    - Audit files are for HUMAN REVIEW ONLY - do not pass to other functions
    - All files are timestamped for uniqueness
    - Low confidence consolidations (< 0.9) are automatically flagged for human review
    
    Args:
        input_csv_path: Path to input CSV file. Must contain columns: business_rule and business_term
                       (column names are case-insensitive and flexible - can be 'Business Rule', 
                       'businessrule', etc.)
        output_dir: Directory where output files will be saved. Defaults to './output'
        generate_excel: Whether to generate formatted Excel output with highlighting for low-confidence
                       consolidations. Defaults to True.
    
    Returns:
        Dictionary with:
        - status: 'success' or 'error'
        - message: Human-readable summary
        - files_created: List of file objects (first is always PRIMARY OUTPUT master contract)
          Each file has: path (full path), filename, type, purpose, row_count, preview
        - summary: Statistics dict with total_original_rules, total_master_rules, rules_consolidated,
          low_confidence_consolidations, business_terms_processed
        - warnings: List of non-fatal warnings
        - errors: List of partial errors (if any business terms failed to process)
        - error_type: (on error) Type of error encountered
        - suggestion: (on error) Suggestion for resolving the error
    """
    
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Step 1: Load and validate
        df_clean, warnings = load_and_validate_contract(input_csv_path)
        
        # Step 2: Extract business terms
        term_categories = extract_business_terms(df_clean)
        
        # Step 3: Process all business terms
        all_results = process_all_business_terms(df_clean, term_categories)
        
        # Step 4: Create master contract
        master_df, audit_df, processing_errors = create_master_contract(all_results)
        
        if master_df.empty:
            output = ConsolidateContractOutput(
                status='error',
                message='Failed to create master contract - no valid rules processed',
                errors=processing_errors,
                warnings=warnings,
                error_type='ProcessingError',
                suggestion='Check that input file contains valid business rules and terms'
            )
            return output.model_dump()
        
        # Step 5: Save outputs
        files_created = save_outputs(master_df, audit_df, output_dir, timestamp)
        
        # Convert file dicts to FileInfo objects
        file_info_objects = [FileInfo(**file_dict) for file_dict in files_created]
        
        # Generate summary
        summary = ConsolidationSummary(
            total_original_rules=len(df_clean),
            total_master_rules=len(master_df),
            rules_consolidated=len(df_clean) - len(master_df),
            low_confidence_consolidations=len(audit_df[audit_df['needs_review'] == True]) if not audit_df.empty else 0,
            business_terms_processed=len(master_df['business_term'].unique()) if not master_df.empty else 0,
            multi_rule_terms_processed=len(term_categories['multi_rule_terms']),
            single_rule_terms_processed=len(term_categories['single_rule_terms'])
        )
        
        # Construct success message
        message = (
            f"Successfully consolidated contract. "
            f"Processed {summary.total_original_rules} original rules into "
            f"{summary.total_master_rules} unique master rules. "
            f"Consolidated {summary.rules_consolidated} duplicate rules across "
            f"{summary.business_terms_processed} business terms."
        )
        
        if summary.low_confidence_consolidations > 0:
            message += f" Note: {summary.low_confidence_consolidations} consolidations flagged for human review (confidence < 0.9)."
        
        output = ConsolidateContractOutput(
            status='success',
            message=message,
            files_created=file_info_objects,
            summary=summary,
            warnings=warnings,
            errors=processing_errors if processing_errors else []
        )
        return output.model_dump()
        
    except FileNotFoundError as e:
        output = ConsolidateContractOutput(
            status='error',
            message=f"File not found: {str(e)}",
            error_type='FileNotFoundError',
            suggestion='Verify the input file path is correct and the file exists.'
        )
        return output.model_dump()
        
    except ValueError as e:
        output = ConsolidateContractOutput(
            status='error',
            message=f"Validation error: {str(e)}",
            error_type='ValueError',
            suggestion='Check that the input CSV has the required columns (business_rule, business_term) and contains valid data.'
        )
        return output.model_dump()
        
    except Exception as e:
        output = ConsolidateContractOutput(
            status='error',
            message=f"Unexpected error during consolidation: {str(e)}",
            error_type=type(e).__name__,
            suggestion='Review the error message and input data. Contact support if the issue persists.'
        )
        return output.model_dump()

# =============================================================================
# END OF TOOL - Schema auto-generated by @tool decorator from Pydantic models
# =============================================================================


# =============================================================================
# USAGE EXAMPLES
# =============================================================================

"""
HOW TO USE THIS TOOL IN LANGGRAPH:

1. Import the tool:
   from consolidate_contract_tool import consolidate_contract

2. Add to your agent's tools list:
   tools = [consolidate_contract, other_tool_1, other_tool_2]

3. The agent can invoke it:
   result = consolidate_contract.invoke({
       "input_csv_path": "consumer_contract.csv",
       "output_dir": "./output"
   })

4. Access structured results (returns dict):
   if result['status'] == "success":
       # First file is always PRIMARY OUTPUT
       master_file_path = result['files_created'][0]['path']
       print(f"Master contract created at: {master_file_path}")
       print(f"Preview:\\n{result['files_created'][0]['preview']}")
       print(f"Consolidated {result['summary']['rules_consolidated']} duplicates")
       
       # Audit files are for human review only
       for file_info in result['files_created'][1:]:
           if "AUDIT ONLY" in file_info['purpose']:
               print(f"Audit trail: {file_info['path']} (do not use as input to other tools)")

IMPORTANT NOTES FOR AGENT BEHAVIOR:
- Only use the master_contract file (first in files_created) as input to other functions
- Audit files are marked with purpose="AUDIT ONLY" - never pass these to other tools
- If status == "error", check error_type and suggestion for recovery steps
- Warnings are non-fatal - process can continue with warnings present
- Files are timestamped (YYYYMMDD_HHMMSS) for uniqueness
- All file paths are FULL PATHS ready to use in subsequent tool calls
"""


# =============================================================================
# FOR STANDALONE TESTING (NOT PART OF TOOL)
# =============================================================================

if __name__ == "__main__":
    # Test the tool standalone
    result = consolidate_contract.invoke({
        "input_csv_path": "test_contract.csv",
        "output_dir": "./output"
    })
    
    print(f"\nStatus: {result['status']}")
    print(f"Message: {result['message']}")
    
    if result['status'] == "success":
        print(f"\nFiles created:")
        for file_info in result['files_created']:
            print(f"  - {file_info['filename']} ({file_info['type']})")
            print(f"    Purpose: {file_info['purpose']}")
            print(f"    Rows: {file_info['row_count']}")
        
        print(f"\nSummary:")
        print(f"  Original rules: {result['summary']['total_original_rules']}")
        print(f"  Master rules: {result['summary']['total_master_rules']}")
        print(f"  Consolidated: {result['summary']['rules_consolidated']}")
    else:
        print(f"\nError: {result['error_type']}")
        print(f"Suggestion: {result['suggestion']}")