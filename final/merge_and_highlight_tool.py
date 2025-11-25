"""
LANGGRAPH TOOL: Merge and Highlight
Merge master contract with new rules delta and highlight for human review

This tool takes the master contract and the new rules identified by compare_contracts,
merges them together, and creates highlighted outputs for human review and approval.
"""

import pandas as pd
import os
from typing import Dict, List, Any, Optional
from datetime import datetime
from pydantic import BaseModel, Field
from langchain_core.tools import tool


# =============================================================================
# CONFIGURATION
# =============================================================================

# Column name variations we should recognize (case-insensitive matching)
COLUMN_ALIASES = {
    'business_rule': ['business_rule', 'business rule', 'businessrule', 'rule', 'business_rules'],
    'business_term': ['business_term', 'business term', 'businessterm', 'term', 'business_terms'],
    'description': ['description', 'desc', 'descriptions'],
    'status': ['status', 'rule_status', 'rule_type']
}


# =============================================================================
# PYDANTIC MODELS FOR TOOL INPUT/OUTPUT
# =============================================================================

class FileInfo(BaseModel):
    """Information about a generated file"""
    path: str = Field(description="Full path to the file")
    filename: str = Field(description="Just the filename with extension")
    type: str = Field(description="Type of file: merged_contract_csv or merged_contract_excel")
    purpose: str = Field(description="PRIMARY OUTPUT - For human review and approval")
    row_count: int = Field(description="Number of rows in the file")
    new_rules_count: int = Field(description="Number of new rules highlighted")
    conflicts_count: int = Field(description="Number of conflicts highlighted in red")
    preview: str = Field(description="Text preview of first 10 rows")


class MergeSummary(BaseModel):
    """Statistics about the merged contract"""
    total_rules: int = Field(description="Total rules in merged contract")
    existing_rules: int = Field(description="Rules from original master contract")
    new_rules_added: int = Field(description="New rules added from delta")
    conflicts_flagged: int = Field(description="Number of conflict rules requiring immediate attention")
    new_stricter_rules: int = Field(description="Number of stricter requirement rules")
    new_different_rules: int = Field(description="Number of different constraint rules")
    new_business_terms: int = Field(description="Number of brand new business terms introduced")


class MergeAndHighlightOutput(BaseModel):
    """Output from the merge_and_highlight tool"""
    status: str = Field(description="'success' or 'error'")
    message: str = Field(description="Human-readable summary of what happened")
    files_created: Optional[List[FileInfo]] = Field(
        default=None,
        description="List of files created. All files are PRIMARY OUTPUT for human review."
    )
    summary: Optional[MergeSummary] = Field(
        default=None,
        description="Detailed statistics about merged contract"
    )
    warnings: List[str] = Field(
        default_factory=list,
        description="Non-fatal warnings encountered during processing"
    )
    error_type: Optional[str] = Field(
        default=None,
        description="Type of error if status is 'error'"
    )
    suggestion: Optional[str] = Field(
        default=None,
        description="Suggestion for resolving the error"
    )


# =============================================================================
# UTILITY: FLEXIBLE COLUMN MATCHING
# =============================================================================

def find_column_name(df: pd.DataFrame, target_column: str) -> Optional[str]:
    """Find the actual column name in DataFrame that matches our target"""
    if target_column not in COLUMN_ALIASES:
        return target_column if target_column in df.columns else None
    
    possible_names = COLUMN_ALIASES[target_column]
    df_columns_lower = {col.lower().strip(): col for col in df.columns}
    
    for possible_name in possible_names:
        possible_lower = possible_name.lower().strip()
        if possible_lower in df_columns_lower:
            return df_columns_lower[possible_lower]
    
    return None


def standardize_dataframe_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Rename DataFrame columns to standard names"""
    rename_map = {}
    
    for standard_name in ['business_rule', 'business_term', 'description', 'status']:
        actual_name = find_column_name(df, standard_name)
        if actual_name and actual_name != standard_name:
            rename_map[actual_name] = standard_name
    
    if rename_map:
        df = df.rename(columns=rename_map)
    
    return df


# =============================================================================
# FILE LOADING
# =============================================================================

def load_master_contract(csv_path: str) -> tuple[pd.DataFrame, List[str]]:
    """
    Load the master contract CSV
    
    Parameters:
        csv_path: Path to master contract CSV
        
    Returns:
        Tuple of (dataframe, warnings_list)
    """
    warnings = []
    
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"Master contract not found: {csv_path}")
    
    df = pd.read_csv(csv_path)
    
    # Standardize column names
    df = standardize_dataframe_columns(df)
    
    # Check required columns
    required_columns = ['business_rule', 'business_term']
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"Master contract missing required columns: {missing_columns}. "
                        f"Available columns: {list(df.columns)}")
    
    if len(df) == 0:
        raise ValueError("Master contract is empty")
    
    return df, warnings


def load_new_rules_delta(csv_path: str) -> tuple[pd.DataFrame, List[str]]:
    """
    Load the new rules delta CSV from compare_contracts
    
    Parameters:
        csv_path: Path to new_rules_delta.csv
        
    Returns:
        Tuple of (dataframe, warnings_list)
    """
    warnings = []
    
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"New rules delta file not found: {csv_path}")
    
    df = pd.read_csv(csv_path)
    
    # Standardize column names
    df = standardize_dataframe_columns(df)
    
    # Check required columns
    required_columns = ['business_rule', 'business_term']
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"New rules delta missing required columns: {missing_columns}. "
                        f"Available columns: {list(df.columns)}")
    
    if len(df) == 0:
        warnings.append("New rules delta is empty - no new rules to merge")
    
    return df, warnings


# =============================================================================
# MERGE LOGIC
# =============================================================================

def merge_contracts(master_df: pd.DataFrame, new_rules_df: pd.DataFrame) -> pd.DataFrame:
    """
    Merge master contract with new rules delta
    
    Parameters:
        master_df: Master contract DataFrame
        new_rules_df: New rules DataFrame from compare_contracts
        
    Returns:
        Merged DataFrame with is_new_rule, status, and other metadata columns
    """
    # Prepare master contract - add new rule indicators
    master_prepared = master_df.copy()
    master_prepared['is_new_rule'] = False
    master_prepared['status'] = 'existing'
    
    # Prepare new rules - select core columns plus metadata
    new_rules_prepared = pd.DataFrame()
    new_rules_prepared['business_rule'] = new_rules_df['business_rule']
    new_rules_prepared['business_term'] = new_rules_df['business_term']
    new_rules_prepared['is_new_rule'] = True
    
    # Add status if it exists
    if 'status' in new_rules_df.columns:
        new_rules_prepared['status'] = new_rules_df['status']
    else:
        new_rules_prepared['status'] = 'new'
    
    # Add other useful metadata columns if they exist
    optional_columns = ['term_type', 'confidence', 'needs_review', 'reasoning', 
                       'related_master_rule', 'original_rule_count']
    for col in optional_columns:
        if col in new_rules_df.columns:
            new_rules_prepared[col] = new_rules_df[col]
    
    # Merge
    merged_df = pd.concat([master_prepared, new_rules_prepared], ignore_index=True)
    
    # Sort by business term, then by is_new_rule (existing rules first, then new rules)
    merged_df = merged_df.sort_values(
        ['business_term', 'is_new_rule'],
        ascending=[True, False]
    ).reset_index(drop=True)
    
    return merged_df


# =============================================================================
# FILE PREVIEW
# =============================================================================

def generate_file_preview(df: pd.DataFrame, max_rows: int = 10) -> str:
    """Generate a text preview of the first N rows of a dataframe"""
    if df.empty:
        return "Empty DataFrame"
    preview_df = df.head(max_rows)
    return preview_df.to_string(index=False)


# =============================================================================
# OUTPUT GENERATION
# =============================================================================

def save_merged_contract(merged_df: pd.DataFrame, 
                        output_dir: str, 
                        timestamp: str,
                        highlight_conflicts_red: bool = True) -> List[Dict[str, Any]]:
    """
    Save merged contract as CSV and Excel with highlighting
    
    Parameters:
        merged_df: Merged DataFrame with is_new_rule and status columns
        output_dir: Output directory path
        timestamp: Timestamp string for unique filenames
        highlight_conflicts_red: Use red highlighting for conflicts (vs yellow)
        
    Returns:
        List of file information dicts
    """
    os.makedirs(output_dir, exist_ok=True)
    
    files_created = []
    
    # Count statistics
    new_rules_count = len(merged_df[merged_df['is_new_rule'] == True])
    conflicts_count = len(merged_df[merged_df['status'] == 'conflict']) if 'status' in merged_df.columns else 0
    
    # Save CSV
    csv_filename = f'master_contract_merged_{timestamp}.csv'
    csv_path = os.path.join(output_dir, csv_filename)
    merged_df.to_csv(csv_path, index=False)
    
    files_created.append({
        'path': csv_path,
        'filename': csv_filename,
        'type': 'merged_contract_csv',
        'purpose': 'PRIMARY OUTPUT - Merged contract for human review and approval',
        'row_count': len(merged_df),
        'new_rules_count': new_rules_count,
        'conflicts_count': conflicts_count,
        'preview': generate_file_preview(merged_df, max_rows=10)
    })
    
    # Try to save Excel with highlighting
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font
        
        excel_filename = f'master_contract_merged_{timestamp}.xlsx'
        excel_path = os.path.join(output_dir, excel_filename)
        
        # Save to Excel first
        merged_df.to_excel(excel_path, index=False, engine='openpyxl')
        
        # Load workbook and apply highlighting
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Define fill colors
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        bold_font = Font(bold=True)
        
        # Find is_new_rule and status columns
        is_new_col = None
        status_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'is_new_rule':
                is_new_col = idx
            if cell.value == 'status':
                status_col = idx
        
        # Apply highlighting to new rules
        if is_new_col:
            red_count = 0
            yellow_count = 0
            
            for row in range(2, ws.max_row + 1):
                is_new = ws.cell(row, is_new_col).value
                
                if is_new == True:
                    # Determine if this is a conflict
                    is_conflict = False
                    if status_col and highlight_conflicts_red:
                        status = ws.cell(row, status_col).value
                        if status == 'conflict':
                            is_conflict = True
                    
                    # Apply appropriate highlighting
                    if is_conflict:
                        # RED for conflicts
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row, col).fill = red_fill
                            ws.cell(row, col).font = bold_font
                        red_count += 1
                    else:
                        # YELLOW for other new rules
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row, col).fill = yellow_fill
                        yellow_count += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(excel_path)
        
        files_created.append({
            'path': excel_path,
            'filename': excel_filename,
            'type': 'merged_contract_excel',
            'purpose': 'PRIMARY OUTPUT - Excel with highlighting (RED=conflicts, YELLOW=new rules)',
            'row_count': len(merged_df),
            'new_rules_count': new_rules_count,
            'conflicts_count': conflicts_count,
            'preview': f'Excel file with {conflicts_count} rows in RED (conflicts) and {new_rules_count - conflicts_count} rows in YELLOW (new rules)'
        })
        
    except ImportError:
        pass  # openpyxl not available, skip Excel
    
    return files_created


# =============================================================================
# MAIN TOOL FUNCTION - LANGGRAPH @tool DECORATOR
# =============================================================================

@tool
def merge_and_highlight(
    master_contract_path: str,
    new_rules_delta_path: str,
    output_dir: str = "./output_final",
    highlight_conflicts_red: bool = True
) -> dict:
    """
    Merge master contract with new rules delta and highlight for human review.
    
    Takes the master/golden contract and the new rules identified by compare_contracts,
    merges them together, and creates highlighted outputs for human review and approval.
    Conflicts are highlighted in RED, other new rules in YELLOW.
    
    IMPORTANT:
    - This is the FINAL step in the contract workflow
    - Output files are for HUMAN REVIEW AND APPROVAL
    - RED highlighting indicates conflicts requiring immediate attention
    - YELLOW highlighting indicates new rules to review
    - Do not pass output files to other automated tools
    
    Args:
        master_contract_path: Full path to master/golden contract CSV file. Must contain 
                            columns: business_rule and business_term (case-insensitive).
        new_rules_delta_path: Full path to new rules delta CSV file (output from compare_contracts).
                             Must contain columns: business_rule, business_term, and optionally
                             status, reasoning, confidence, etc.
        output_dir: Directory where output files will be saved. Defaults to './output_final'
        highlight_conflicts_red: Whether to use red highlighting for conflicts vs yellow.
                                Defaults to True (conflicts in red, other new rules in yellow).
    
    Returns:
        Dictionary with:
        - status: 'success' or 'error'
        - message: Human-readable summary
        - files_created: List of file objects (all are PRIMARY OUTPUT for human review)
          Each file has: path (full path), filename, type, purpose, row_count, 
          new_rules_count, conflicts_count, preview
        - summary: Statistics dict with total_rules, existing_rules, new_rules_added,
          conflicts_flagged, new_stricter_rules, new_different_rules, new_business_terms
        - warnings: List of non-fatal warnings
        - error_type: (on error) Type of error encountered
        - suggestion: (on error) Suggestion for resolving the error
    """
    
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        all_warnings = []
        
        # Step 1: Load master contract
        master_df, master_warnings = load_master_contract(master_contract_path)
        all_warnings.extend(master_warnings)
        
        # Step 2: Load new rules delta
        new_rules_df, delta_warnings = load_new_rules_delta(new_rules_delta_path)
        all_warnings.extend(delta_warnings)
        
        # Check if new rules delta is empty
        if len(new_rules_df) == 0:
            output = MergeAndHighlightOutput(
                status='success',
                message='No new rules to merge. The proposed contract is fully covered by the master contract.',
                files_created=[],
                summary=MergeSummary(
                    total_rules=len(master_df),
                    existing_rules=len(master_df),
                    new_rules_added=0,
                    conflicts_flagged=0,
                    new_stricter_rules=0,
                    new_different_rules=0,
                    new_business_terms=0
                ),
                warnings=all_warnings
            )
            return output.model_dump()
        
        # Step 3: Merge contracts
        merged_df = merge_contracts(master_df, new_rules_df)
        
        # Step 4: Save outputs
        files_created = save_merged_contract(merged_df, output_dir, timestamp, highlight_conflicts_red)
        
        # Convert file dicts to FileInfo objects
        file_info_objects = [FileInfo(**file_dict) for file_dict in files_created]
        
        # Step 5: Generate summary statistics
        new_rules_count = len(merged_df[merged_df['is_new_rule'] == True])
        
        # Count by status if available
        conflicts_count = 0
        new_stricter_count = 0
        new_different_count = 0
        new_term_count = 0
        
        if 'status' in merged_df.columns:
            status_counts = merged_df[merged_df['is_new_rule'] == True]['status'].value_counts()
            conflicts_count = status_counts.get('conflict', 0)
            new_stricter_count = status_counts.get('new_stricter', 0)
            new_different_count = status_counts.get('new_different', 0)
            new_term_count = status_counts.get('unique', 0)
        
        summary = MergeSummary(
            total_rules=len(merged_df),
            existing_rules=len(master_df),
            new_rules_added=new_rules_count,
            conflicts_flagged=conflicts_count,
            new_stricter_rules=new_stricter_count,
            new_different_rules=new_different_count,
            new_business_terms=new_term_count
        )
        
        # Construct success message
        message = (
            f"Successfully merged contracts. Created merged contract with {summary.total_rules} total rules: "
            f"{summary.existing_rules} existing rules + {summary.new_rules_added} new rules."
        )
        
        if conflicts_count > 0:
            message += f" CRITICAL: {conflicts_count} conflicts highlighted in RED require immediate review."
        
        if new_rules_count - conflicts_count > 0:
            message += f" {new_rules_count - conflicts_count} new rules highlighted in YELLOW for approval."
        
        output = MergeAndHighlightOutput(
            status='success',
            message=message,
            files_created=file_info_objects,
            summary=summary,
            warnings=all_warnings
        )
        return output.model_dump()
        
    except FileNotFoundError as e:
        output = MergeAndHighlightOutput(
            status='error',
            message=f"File not found: {str(e)}",
            error_type='FileNotFoundError',
            suggestion='Verify both file paths are correct and files exist. Ensure compare_contracts was run successfully.'
        )
        return output.model_dump()
        
    except ValueError as e:
        output = MergeAndHighlightOutput(
            status='error',
            message=f"Validation error: {str(e)}",
            error_type='ValueError',
            suggestion='Check that both CSV files have required columns (business_rule, business_term) and contain valid data.'
        )
        return output.model_dump()
        
    except Exception as e:
        output = MergeAndHighlightOutput(
            status='error',
            message=f"Unexpected error during merge: {str(e)}",
            error_type=type(e).__name__,
            suggestion='Review the error message and input data. Contact support if the issue persists.'
        )
        return output.model_dump()


# =============================================================================
# END OF TOOL - Schema auto-generated by @tool decorator from Pydantic models
# =============================================================================


# =============================================================================
# USAGE EXAMPLES
# =============================================================================

"""
HOW TO USE THIS TOOL IN LANGGRAPH:

1. Import the tool:
   from merge_and_highlight_tool import merge_and_highlight

2. Add to your agent's tools list:
   tools = [consolidate_contract, compare_contracts, merge_and_highlight]

3. The agent can invoke it (typically after compare_contracts):
   # Complete workflow:
   result1 = consolidate_contract.invoke({"input_csv_path": "consumer.csv"})
   master_path = result1['files_created'][0]['path']
   
   result2 = compare_contracts.invoke({
       "master_contract_path": master_path,
       "proposed_contract_path": "proposed.csv"
   })
   
   # Only merge if there are new rules
   if result2['summary']['new_rules_identified'] > 0:
       delta_path = result2['files_created'][0]['path']
       
       result3 = merge_and_highlight.invoke({
           "master_contract_path": master_path,
           "new_rules_delta_path": delta_path
       })

4. Access structured results (returns dict):
   if result3['status'] == "success":
       # All files are for human review
       for file_info in result3['files_created']:
           print(f"Review file: {file_info['path']}")
           if file_info['type'] == 'merged_contract_excel':
               print(f"  Conflicts (RED): {file_info['conflicts_count']}")
               print(f"  New rules (YELLOW): {file_info['new_rules_count']}")
       
       # Check for conflicts requiring immediate attention
       if result3['summary']['conflicts_flagged'] > 0:
           print(f"CRITICAL: {result3['summary']['conflicts_flagged']} conflicts need resolution!")

IMPORTANT NOTES FOR AGENT BEHAVIOR:
- This is the FINAL step - output is for HUMAN REVIEW, not automated processing
- Do NOT pass output files to other tools
- RED highlighting = conflicts requiring immediate human attention
- YELLOW highlighting = new rules needing approval
- If new_rules_delta is empty, no merge is performed
- Agent should inform user to review highlighted Excel file
- All file paths are FULL PATHS
"""


# =============================================================================
# FOR STANDALONE TESTING (NOT PART OF TOOL)
# =============================================================================

if __name__ == "__main__":
    # Test the tool standalone
    result = merge_and_highlight.invoke({
        "master_contract_path": "output/master_contract_20250524.csv",
        "new_rules_delta_path": "output/new_rules_delta_20250524.csv"
    })
    
    print(f"\nStatus: {result['status']}")
    print(f"Message: {result['message']}")
    
    if result['status'] == "success":
        print(f"\nFiles created:")
        for file_info in result['files_created']:
            print(f"  - {file_info['filename']} ({file_info['type']})")
            print(f"    Purpose: {file_info['purpose']}")
            print(f"    Total rows: {file_info['row_count']}")
            print(f"    New rules: {file_info['new_rules_count']}")
            print(f"    Conflicts: {file_info['conflicts_count']}")
        
        print(f"\nSummary:")
        print(f"  Total rules: {result['summary']['total_rules']}")
        print(f"  Existing: {result['summary']['existing_rules']}")
        print(f"  New added: {result['summary']['new_rules_added']}")
        print(f"  Conflicts: {result['summary']['conflicts_flagged']}")
    else:
        print(f"\nError: {result['error_type']}")
        print(f"Suggestion: {result['suggestion']}")