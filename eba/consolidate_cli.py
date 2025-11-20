"""
CONTRACT CONSOLIDATION POC
Simple, readable implementation for client demonstration
"""

import boto3
import json
import pandas as pd
import os
import argparse
from typing import Dict, List, Any


# =============================================================================
# CONFIGURATION
# =============================================================================

MAX_RETRIES = 3


# =============================================================================
# CLI ARGUMENT PARSING
# =============================================================================

def parse_arguments():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(
        description='Contract Consolidation POC - Consolidate duplicate business rules using LLM',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Basic usage with default settings
  python contract_consolidator.py --input contracts.csv
  
  # Specify custom output directory
  python contract_consolidator.py --input contracts.csv --output results/
  
  # Use different AWS region and model
  python contract_consolidator.py --input contracts.csv --region us-east-1 --model claude-haiku-4-20250514
  
  # Adjust confidence threshold for review flagging
  python contract_consolidator.py --input contracts.csv --confidence-threshold 0.85
        """
    )
    
    parser.add_argument(
        '--input', '-i',
        type=str,
        required=True,
        help='Path to input CSV file containing business rules (required columns: Business Rule, Business Term)'
    )
    
    parser.add_argument(
        '--output', '-o',
        type=str,
        default='./output',
        help='Output directory for results (default: ./output)'
    )
    
    parser.add_argument(
        '--region', '-r',
        type=str,
        default='us-west-2',
        help='AWS Bedrock region (default: us-west-2)'
    )
    
    parser.add_argument(
        '--model', '-m',
        type=str,
        default='us.anthropic.claude-sonnet-4-20250514',
        help='Bedrock model ID (default: claude-sonnet-4-20250514)'
    )
    
    parser.add_argument(
        '--temperature', '-t',
        type=float,
        default=0.0,
        help='Model temperature (default: 0.0 for consistency)'
    )
    
    parser.add_argument(
        '--max-tokens',
        type=int,
        default=4096,
        help='Maximum tokens per LLM response (default: 4096)'
    )
    
    parser.add_argument(
        '--max-retries',
        type=int,
        default=3,
        help='Maximum retry attempts for failed LLM calls (default: 3)'
    )
    
    parser.add_argument(
        '--confidence-threshold',
        type=float,
        default=0.9,
        help='Confidence threshold for flagging consolidations for review (default: 0.9)'
    )
    
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Enable verbose output'
    )
    
    return parser.parse_args()


# =============================================================================
# SYSTEM PROMPT
# =============================================================================

def get_system_prompt():
    """Returns the system prompt for rule consolidation"""
    return """You are a data quality expert analyzing business rules to identify and eliminate true duplicates. Your goal is to ensure each unique business logic constraint is preserved while removing redundant rules that say the same thing in different words.

## INPUT FORMAT

Business rules are formatted as "[Business Term]: [Rule Description]"

**CRITICAL**: If a rule starts with text that matches or is similar to the business term, ignore that prefix. Only analyze the actual rule logic after the colon.

Examples:
- "Account Category: The value must be chosen from an approved list" → Analyze: "The value must be chosen from an approved list"
- "FI - Muni/State Invested: This field is required" → Analyze: "This field is required"

## WHAT MAKES RULES DUPLICATES

Two rules are duplicates ONLY if they enforce the exact same business logic constraint. They must be identical in:
1. **What** is being constrained
2. **How** it is constrained  
3. **When** it applies (conditions/scope)

### DUPLICATES - Same constraint, different wording:
- "Field is required" ≡ "Field cannot be empty" ≡ "Field must not be null" ≡ "This field is required and cannot be empty"
- "Value must be from approved list" ≡ "Value must be chosen from a standard list" ≡ "The value must be chosen from an approved list"
- "CUSIP must be 9 characters" ≡ "CUSIP length must equal 9"

### NOT DUPLICATES - Different business logic:
- "Field is required" ≠ "Field is required only if X is Y" (unconditional vs conditional)
- "Field is required if X is Y" ≠ "Field is required if X is Z" (different conditions)
- "Value from approved list" ≠ "Field cannot be empty" (different constraints)
- "Value from list if OMS is AIM" ≠ "Value from list if OMS is BBGAIM" (different condition values)
- "Field > 0" ≠ "Field >= 0" (different boundaries)

## CONDITIONAL LOGIC IS CRITICAL

Rules with conditional qualifiers ("if", "when", "only if", "only when") are fundamentally different from unconditional rules.

**IMPORTANT**: A conditional rule is NEVER a duplicate of an unconditional rule, even if the constraint text is similar.

Examples of UNIQUE rules (not duplicates):
- "This field is required"
- "This field is required only if the Muni Inv Focus is state"

These define different business logic - one always applies, one conditionally applies.

## COMPLEMENTARY RULES

Multiple rules for the same business term may represent different aspects of validation. These are NOT duplicates:

Example - All three are UNIQUE:
- "If the OMS designation is AIM, this field cannot be null"
- "The value must be chosen from an approved list"
- "This field is required and cannot be empty"

These rules complement each other to form a complete validation contract. Keep them all unless they're truly redundant.

## RULES WITH AND/OR OPERATORS

If a rule contains AND/OR operators, focus on the complete business logic being expressed:
- "Cannot be negative and cannot be zero" might equal "Must be greater than 0"
- "Status is active or pending" equals "Status must be one of: active, pending"

Compare the complete logical constraint, not individual parts.

## CANONICAL RULE SELECTION

When rules ARE duplicates, select the clearest version as canonical:

Priority:
1. **Clarity**: Most unambiguous and explicit
2. **Completeness**: Captures full constraint without being verbose
3. **Standard phrasing**: Prefer "must" over "should", positive statements over negative

Example: For duplicates "required", "cannot be null", "must not be empty"
Select: "This field is required and cannot be empty" (most complete and clear)

## CONFIDENCE SCORING

- **1.0**: Absolutely certain (mathematically/logically identical)
- **0.9-0.99**: Very confident (clear semantic equivalence, minor wording differences)
- **0.8-0.89**: Confident (likely equivalent with slight ambiguity)
- **< 0.8**: When uncertain, mark as UNIQUE rather than duplicate

Be conservative. If there's any doubt about whether rules express the same business logic, keep them as unique.

## REASONING REQUIREMENTS

For every rule, explain your decision clearly:
- Why is it unique or a duplicate?
- If duplicate, what makes it semantically identical?
- If unique, what differentiates it (different constraint, different condition, different scope)?

Good: "Unique. This rule is conditional ('only if Muni Inv Focus is state') while rule at index 1 is unconditional. Different application scope means different business logic."

Good: "Duplicate of rule at index 0. Both enforce the 'required' constraint unconditionally. Different phrasing ('cannot be empty' vs 'is required') but identical business logic. Confidence 1.0."

Bad: "These are different" (no explanation)
Bad: "Duplicate" (doesn't explain why)

## OUTPUT FORMAT

Return ONLY valid JSON with no markdown formatting or additional text:

{
    "business_term": "string",
    "total_rules_analyzed": 0,
    "unique_rules_identified": 0,
    "duplicates_found": false,
    "analysis": [
        {
            "original_rule": "exact input text with prefix",
            "original_rule_index": 0,
            "status": "unique|duplicate",
            "canonical_rule": "best version of rule (with prefix removed)",
            "confidence": 0.95,
            "consolidated_with": ["array of other rule texts if duplicate"],
            "duplicate_of_index": null,
            "reasoning": "clear explanation of decision"
        }
    ],
    "notes": "any additional observations"
}

**Field notes:**
- `original_rule`: Include the full text as provided (with prefix)
- `canonical_rule`: Remove the business term prefix, use clearest phrasing
- `status`: Only "unique" or "duplicate"
- `confidence`: 0.0 to 1.0
- `duplicate_of_index`: Index of the canonical rule if this is a duplicate, null if unique
- `reasoning`: Required - explain your logic (2-3 sentences)

Return analysis array in the same order as input rules.

## EDGE CASES

- **Boundary differences**: "Age > 18" vs "Age >= 18" are NOT duplicates
- **Different condition values**: "if X is Y" vs "if X is Z" are NOT duplicates  
- **Subset relationships**: "Age > 21" is stricter than "Age > 18" but both are unique
- **Multi-field references**: Rules referencing different fields are NOT duplicates

When in doubt, preserve uniqueness. It's better to keep a rule that might be duplicate than to incorrectly consolidate rules with different business logic."""


# =============================================================================
# DATA LOADING
# =============================================================================

def load_consumer_contract(csv_path: str) -> pd.DataFrame:
    """
    Load the input CSV file containing business rules
    
    Parameters:
        csv_path: Path to input CSV file
        
    Returns:
        DataFrame with columns: Business Rule, Description, Business Term
    """
    print(f"\n{'='*60}")
    print("STEP 1: Loading contract data")
    print(f"{'='*60}")
    
    df = pd.read_csv(csv_path)
    
    print(f"✓ Loaded {len(df)} rules from {csv_path}")
    print(f"  Columns: {', '.join(df.columns.tolist())}")
    
    return df


# =============================================================================
# DATA VALIDATION AND CLEANING
# =============================================================================

def validate_and_clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Apply validation and quality checks to the dataframe
    
    Parameters:
        df: Raw input data
        
    Returns:
        Cleaned and validated DataFrame
    """
    print(f"\n{'='*60}")
    print("STEP 2: Validating and cleaning data")
    print(f"{'='*60}")
    
    original_count = len(df)
    
    # Check required columns exist
    required_columns = ['Business Rule', 'Business Term']
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"Missing required columns: {missing_columns}")
    print(f"✓ All required columns present")
    
    # Remove rows with null business_rule or business_term
    df = df.dropna(subset=['Business Rule', 'Business Term'])
    null_removed = original_count - len(df)
    if null_removed > 0:
        print(f"✓ Removed {null_removed} rows with null values")
    
    # Trim whitespace
    df['Business Rule'] = df['Business Rule'].str.strip()
    df['Business Term'] = df['Business Term'].str.strip()
    print(f"✓ Trimmed whitespace from text fields")
    
    # Remove exact duplicates
    before_dedup = len(df)
    df = df.drop_duplicates(subset=['Business Rule', 'Business Term'])
    duplicates_removed = before_dedup - len(df)
    if duplicates_removed > 0:
        print(f"✓ Removed {duplicates_removed} exact duplicate rows")
    
    print(f"\n  Final cleaned dataset: {len(df)} rules")
    
    return df


# =============================================================================
# BUSINESS TERM EXTRACTION
# =============================================================================

def extract_unique_business_terms(df: pd.DataFrame) -> List[str]:
    """
    Get list of business terms that need consolidation (2+ rules)
    
    Parameters:
        df: Cleaned data
        
    Returns:
        List of business terms with 2 or more rules
    """
    # Group by business term and count
    term_counts = df.groupby('Business Term').size()
    
    # Filter to terms with 2+ rules
    terms_needing_consolidation = term_counts[term_counts >= 2].index.tolist()
    
    # Sort alphabetically
    terms_needing_consolidation.sort()
    
    return terms_needing_consolidation


def get_all_business_terms(df: pd.DataFrame) -> List[str]:
    """
    Get complete list of all unique business terms
    
    Parameters:
        df: Cleaned data
        
    Returns:
        List of all unique business terms
    """
    return df['Business Term'].unique().tolist()


# =============================================================================
# RULE FILTERING
# =============================================================================

def get_rules_for_business_term(df: pd.DataFrame, business_term: str) -> pd.DataFrame:
    """
    Filter dataframe to only rules for a specific business term
    
    Parameters:
        df: Full cleaned data
        business_term: The business term to filter for
        
    Returns:
        DataFrame subset containing only rows for this business term
    """
    return df[df['Business Term'] == business_term].copy()


# =============================================================================
# LLM INVOCATION
# =============================================================================

def invoke_bedrock_model(system_prompt: str, user_prompt: str, config: dict) -> str:
    """
    Call AWS Bedrock API to invoke Claude
    
    Parameters:
        system_prompt: System instructions
        user_prompt: User message
        config: Configuration dictionary with region, model_id, etc.
        
    Returns:
        Raw response text from model
    """
    # Initialize Bedrock client
    bedrock = boto3.client(
        service_name='bedrock-runtime',
        region_name='us-east-1'
    )
    
    # Build request
    body = {
        "modelId": "us.anthropic.claude-haiku-4-5-20251001-v1:0",
        "inferenceConfig": {
            "maxTokens": 4096,
            "temperature": 0
        },
        "system": [{"text": system_prompt}],
        "messages": [{
            'role': 'user',
            'content': [{'text': user_prompt}]
        }]
    }
    
    # Call API
    response = bedrock.converse(**body)
    
    # Extract text
    result = response['output']['message']['content'][0]['text']
    
    return result


# =============================================================================
# LLM CONSOLIDATION
# =============================================================================

def consolidate_rules_with_llm(rules_df: pd.DataFrame, business_term: str, config: dict) -> Dict[str, Any]:
    """
    Use LLM to semantically consolidate duplicate rules for a business term
    
    Parameters:
        rules_df: All rules for one business term
        business_term: The business term being processed
        config: Configuration dictionary
        
    Returns:
        Dictionary with consolidation analysis
    """
    max_retries = config.get('max_retries', MAX_RETRIES)
    verbose = config.get('verbose', False)
    
    # Extract list of rules
    rules_list = rules_df['Business Rule'].tolist()
    
    # Build user prompt
    user_prompt = f"""Business Term: {business_term}

Rules to analyze:
{json.dumps(rules_list, indent=2)}

Analyze these rules and identify semantic duplicates. Return your analysis as valid JSON only (no markdown, no code blocks).

Required JSON structure:
{{
    "business_term": "{business_term}",
    "total_rules_analyzed": <number>,
    "unique_rules_identified": <number>,
    "duplicates_found": <boolean>,
    "analysis": [
        {{
            "original_rule": "<exact rule text>",
            "original_rule_index": <index>,
            "status": "unique" or "duplicate",
            "canonical_rule": "<best version>",
            "confidence": <0.0 to 1.0>,
            "consolidated_with": [<array of duplicate rule texts>],
            "duplicate_of_index": <index or null>,
            "reasoning": "<clear explanation>"
        }}
    ],
    "notes": "<any observations>"
}}"""
    
    # Get system prompt
    system_prompt = get_system_prompt()
    
    # Attempt with retries
    for attempt in range(max_retries):
        try:
            if verbose or attempt > 0:
                print(f"  → Calling LLM (attempt {attempt + 1}/{max_retries})...")
            else:
                print(f"  → Calling LLM...")
            
            # Call LLM
            llm_response = invoke_bedrock_model(system_prompt, user_prompt, config)
            
            # Try to parse as JSON
            # Remove any markdown code blocks if present
            llm_response = llm_response.strip()
            if llm_response.startswith("```"):
                # Remove markdown code blocks
                lines = llm_response.split('\n')
                llm_response = '\n'.join(lines[1:-1])
            
            result = json.loads(llm_response)
            
            # Validate structure
            validate_llm_response_structure(result, business_term)
            
            print(f"  ✓ Successfully parsed LLM response")
            return result
            
        except (json.JSONDecodeError, ValueError, KeyError) as e:
            print(f"  ✗ Attempt {attempt + 1} failed: {str(e)}")
            
            if attempt < max_retries - 1:
                # Modify prompt to emphasize JSON-only
                user_prompt += "\n\nIMPORTANT: Return ONLY valid JSON. No markdown code blocks, no explanations outside JSON."
            else:
                # All retries failed
                print(f"  ✗ ERROR: Failed to parse LLM response after {max_retries} attempts")
                return {
                    "business_term": business_term,
                    "error": f"Failed to parse LLM response: {str(e)}",
                    "total_rules_analyzed": len(rules_list),
                    "unique_rules_identified": 0,
                    "duplicates_found": False,
                    "analysis": []
                }


def validate_llm_response_structure(response: Dict, business_term: str):
    """
    Validate that LLM response has expected structure
    
    Parameters:
        response: Parsed JSON from LLM
        business_term: Expected business term
        
    Raises:
        ValueError if validation fails
    """
    # Check required fields
    if 'business_term' not in response:
        raise ValueError("Missing 'business_term' field")
    
    if 'analysis' not in response:
        raise ValueError("Missing 'analysis' field")
    
    if not isinstance(response['analysis'], list):
        raise ValueError("'analysis' must be a list")
    
    # Validate each analysis entry
    for i, entry in enumerate(response['analysis']):
        required_fields = ['original_rule', 'status', 'canonical_rule', 'confidence']
        for field in required_fields:
            if field not in entry:
                raise ValueError(f"Analysis entry {i} missing field: {field}")
        
        # Check status value
        if entry['status'] not in ['unique', 'duplicate']:
            raise ValueError(f"Analysis entry {i} has invalid status: {entry['status']}")
        
        # Check confidence range
        if not (0.0 <= entry['confidence'] <= 1.0):
            raise ValueError(f"Analysis entry {i} has invalid confidence: {entry['confidence']}")


# =============================================================================
# SINGLE RULE PROCESSING
# =============================================================================

def create_single_rule_audit_entry(business_term: str, rule_text: str) -> Dict[str, Any]:
    """
    Create audit entry for business terms with only one rule (no LLM needed)
    
    Parameters:
        business_term: The business term
        rule_text: The single rule text
        
    Returns:
        Dictionary matching LLM output format
    """
    return {
        "business_term": business_term,
        "total_rules_analyzed": 1,
        "unique_rules_identified": 1,
        "duplicates_found": False,
        "analysis": [
            {
                "original_rule": rule_text,
                "original_rule_index": 0,
                "status": "unique",
                "canonical_rule": rule_text,
                "confidence": 1.0,
                "consolidated_with": [],
                "duplicate_of_index": None,
                "reasoning": "Only one rule detected for this business term. No consolidation needed."
            }
        ],
        "notes": "Single rule - no LLM processing required"
    }


# =============================================================================
# ORCHESTRATION
# =============================================================================

def process_all_business_terms(df: pd.DataFrame, 
                                terms_needing_consolidation: List[str],
                                all_terms: List[str],
                                config: dict) -> List[Dict[str, Any]]:
    """
    Process all business terms - both multi-rule (with LLM) and single-rule (without LLM)
    
    Parameters:
        df: Cleaned full dataset
        terms_needing_consolidation: Terms with 2+ rules
        all_terms: All business terms
        config: Configuration dictionary
        
    Returns:
        List of consolidation results for all terms
    """
    print(f"\n{'='*60}")
    print("STEP 3: Processing business terms")
    print(f"{'='*60}")
    
    all_results = []
    
    # Process multi-rule terms (needs LLM)
    print(f"\nProcessing {len(terms_needing_consolidation)} terms with multiple rules:")
    for i, business_term in enumerate(terms_needing_consolidation, 1):
        # Get rules for this term
        term_rules_df = get_rules_for_business_term(df, business_term)
        rule_count = len(term_rules_df)
        
        print(f"\n[{i}/{len(terms_needing_consolidation)}] {business_term} ({rule_count} rules)")
        
        # Consolidate with LLM
        result = consolidate_rules_with_llm(term_rules_df, business_term, config)
        all_results.append(result)
    
    # Process single-rule terms (no LLM needed)
    single_rule_terms = [t for t in all_terms if t not in terms_needing_consolidation]
    
    if single_rule_terms:
        print(f"\n\nProcessing {len(single_rule_terms)} terms with single rules (no LLM needed):")
        for business_term in single_rule_terms:
            term_rules_df = get_rules_for_business_term(df, business_term)
            
            if len(term_rules_df) == 1:
                rule_text = term_rules_df.iloc[0]['Business Rule']
                result = create_single_rule_audit_entry(business_term, rule_text)
                all_results.append(result)
                print(f"  • {business_term}")
    
    return all_results


# =============================================================================
# MASTER CONTRACT CREATION
# =============================================================================

def create_master_contract(all_results: List[Dict[str, Any]], config: dict) -> tuple:
    """
    Create master contract CSV and detailed audit CSV from LLM results
    
    Parameters:
        all_results: Consolidation results from all business terms
        config: Configuration dictionary
        
    Returns:
        Tuple of (master_df, audit_df)
    """
    print(f"\n{'='*60}")
    print("STEP 4: Creating master contract and audit trail")
    print(f"{'='*60}")
    
    confidence_threshold = config.get('confidence_threshold', 0.9)
    
    master_rules = []
    audit_records = []
    
    # Process each result
    for result in all_results:
        business_term = result['business_term']
        
        # Skip if error occurred
        if 'error' in result:
            print(f"  ⚠ Warning: Skipping {business_term} due to error: {result['error']}")
            continue
        
        # Track which canonical rules we've seen
        seen_canonical = set()
        
        # Process each rule analysis
        for rule_analysis in result['analysis']:
            # Create audit record
            audit_record = {
                'original_rule': rule_analysis['original_rule'],
                'business_term': business_term,
                'status': rule_analysis['status'],
                'canonical_rule': rule_analysis['canonical_rule'],
                'confidence': rule_analysis['confidence'],
                'needs_review': rule_analysis['confidence'] < confidence_threshold,
                'reasoning': rule_analysis['reasoning']
            }
            audit_records.append(audit_record)
            
            # Add unique rules to master (only once per canonical rule)
            canonical_rule = rule_analysis['canonical_rule']
            if canonical_rule not in seen_canonical:
                seen_canonical.add(canonical_rule)
                
                # Count how many original rules map to this canonical
                original_rule_count = sum(
                    1 for r in result['analysis'] 
                    if r['canonical_rule'] == canonical_rule
                )
                
                # Determine consolidation method
                if len(result['analysis']) == 1:
                    consolidation_method = 'single_rule'
                else:
                    consolidation_method = 'llm_consolidated'
                
                master_rules.append({
                    'business_rule': canonical_rule,
                    'business_term': business_term,
                    'consolidation_method': consolidation_method,
                    'original_rule_count': original_rule_count
                })
    
    # Convert to DataFrames
    master_df = pd.DataFrame(master_rules)
    audit_df = pd.DataFrame(audit_records)
    
    # Sort for consistency
    master_df = master_df.sort_values(['business_term', 'business_rule']).reset_index(drop=True)
    audit_df = audit_df.sort_values(
        ['business_term', 'needs_review', 'original_rule'],
        ascending=[True, False, True]
    ).reset_index(drop=True)
    
    print(f"✓ Created master contract with {len(master_df)} unique rules")
    print(f"✓ Created audit trail with {len(audit_df)} total rules analyzed")
    
    return master_df, audit_df


# =============================================================================
# OUTPUT GENERATION
# =============================================================================

def save_outputs(master_df: pd.DataFrame, audit_df: pd.DataFrame, output_dir: str = './output'):
    """
    Save master contract and audit trail
    
    Parameters:
        master_df: Consolidated master contract
        audit_df: Detailed audit trail
        output_dir: Directory to save outputs
        
    Returns:
        Dictionary with summary statistics
    """
    print(f"\n{'='*60}")
    print("STEP 5: Saving outputs")
    print(f"{'='*60}")
    
    # Create output directory
    os.makedirs(output_dir, exist_ok=True)
    
    # Save master contract
    master_path = os.path.join(output_dir, 'master_contract.csv')
    master_df.to_csv(master_path, index=False)
    print(f"✓ Saved master contract: {master_path}")
    
    # Save audit trail
    audit_path = os.path.join(output_dir, 'consolidation_audit.csv')
    audit_df.to_csv(audit_path, index=False)
    print(f"✓ Saved audit trail: {audit_path}")
    
    # Try to save Excel with formatting
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font
        
        excel_path = os.path.join(output_dir, 'consolidation_audit_formatted.xlsx')
        audit_df.to_excel(excel_path, index=False, engine='openpyxl')
        
        # Apply formatting
        wb = load_workbook(excel_path)
        ws = wb.active
        
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        bold_font = Font(bold=True)
        
        # Find needs_review column
        needs_review_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'needs_review':
                needs_review_col = idx
                break
        
        # Highlight rows where needs_review = TRUE
        if needs_review_col:
            for row in range(2, ws.max_row + 1):
                if ws.cell(row, needs_review_col).value == True:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row, col).fill = yellow_fill
        
        wb.save(excel_path)
        print(f"✓ Saved formatted audit trail: {excel_path}")
        
    except ImportError:
        print(f"  ℹ openpyxl not available - skipping Excel export")
        print(f"    Install with: pip install openpyxl")
    
    # Generate summary statistics
    summary = {
        'total_original_rules': len(audit_df),
        'total_master_rules': len(master_df),
        'rules_consolidated': len(audit_df) - len(master_df),
        'low_confidence_consolidations': len(audit_df[audit_df['needs_review'] == True]),
        'business_terms_processed': len(master_df['business_term'].unique())
    }
    
    # Save summary
    summary_path = os.path.join(output_dir, 'consolidation_summary.json')
    with open(summary_path, 'w') as f:
        json.dump(summary, f, indent=2)
    print(f"✓ Saved summary: {summary_path}")
    
    return summary


# =============================================================================
# MAIN WORKFLOW
# =============================================================================

def main():
    """Execute the complete contract consolidation workflow"""
    
    # Parse command line arguments
    args = parse_arguments()
    
    # Build configuration dictionary
    config = {
        'region': args.region,
        'model_id': args.model,
        'temperature': args.temperature,
        'max_tokens': args.max_tokens,
        'max_retries': args.max_retries,
        'confidence_threshold': args.confidence_threshold,
        'verbose': args.verbose
    }
    
    print("\n" + "="*60)
    print("CONTRACT CONSOLIDATION POC")
    print("="*60)
    
    if args.verbose:
        print("\nConfiguration:")
        print(f"  Input file: {args.input}")
        print(f"  Output directory: {args.output}")
        print(f"  AWS Region: {config['region']}")
        print(f"  Model: {config['model_id']}")
        print(f"  Temperature: {config['temperature']}")
        print(f"  Max tokens: {config['max_tokens']}")
        print(f"  Max retries: {config['max_retries']}")
        print(f"  Confidence threshold: {config['confidence_threshold']}")
    
    # STEP 1: Load data
    df = load_consumer_contract(args.input)
    
    # STEP 2: Validate and clean
    df_clean = validate_and_clean_dataframe(df)
    
    # Extract business terms
    all_business_terms = get_all_business_terms(df_clean)
    terms_needing_consolidation = extract_unique_business_terms(df_clean)
    
    print(f"\nProcessing Summary:")
    print(f"  • Total business terms: {len(all_business_terms)}")
    print(f"  • Terms needing consolidation (2+ rules): {len(terms_needing_consolidation)}")
    print(f"  • Single-rule terms (audit only): {len(all_business_terms) - len(terms_needing_consolidation)}")
    
    # STEP 3: Process all business terms
    all_results = process_all_business_terms(
        df_clean,
        terms_needing_consolidation,
        all_business_terms,
        config
    )
    
    # STEP 4: Create master contract
    master_df, audit_df = create_master_contract(all_results, config)
    
    # STEP 5: Save outputs
    summary = save_outputs(master_df, audit_df, args.output)
    
    # Final summary
    print(f"\n{'='*60}")
    print("CONSOLIDATION COMPLETE")
    print(f"{'='*60}")
    print(f"\nResults:")
    print(f"  • Original rules: {len(df)}")
    print(f"  • After cleaning: {len(df_clean)}")
    print(f"  • Master contract rules: {summary['total_master_rules']}")
    print(f"  • Rules consolidated: {summary['rules_consolidated']}")
    print(f"  • Reduction: {summary['rules_consolidated'] / len(df_clean) * 100:.1f}%")
    
    if summary['low_confidence_consolidations'] > 0:
        print(f"\n⚠️  {summary['low_confidence_consolidations']} consolidations flagged for review (confidence < {config['confidence_threshold']})")
        print(f"   Review in: {args.output}/consolidation_audit_formatted.xlsx (highlighted rows)")
        print(f"   Or filter: {args.output}/consolidation_audit.csv where needs_review=TRUE")
    
    print(f"\n✓ All outputs saved to {args.output}/")
    print("="*60 + "\n")


if __name__ == "__main__":
    main()