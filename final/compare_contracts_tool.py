"""
LANGGRAPH TOOL: Compare Contracts
Compare proposed contract against master contract to identify new rules (delta detection)

This tool performs semantic comparison to identify which rules in a proposed contract
are genuinely new versus already covered by the master contract.
"""

import boto3
import json
import pandas as pd
import os
from typing import Dict, List, Any, Optional
from datetime import datetime
from pydantic import BaseModel, Field
from langchain_core.tools import tool


# =============================================================================
# CONFIGURATION
# =============================================================================

MAX_RETRIES = 3

# Column name variations we should recognize (case-insensitive matching)
COLUMN_ALIASES = {
    'business_rule': ['business_rule', 'business rule', 'businessrule', 'rule', 'business_rules'],
    'business_term': ['business_term', 'business term', 'businessterm', 'term', 'business_terms'],
    'description': ['description', 'desc', 'descriptions']
}


# =============================================================================
# PYDANTIC MODELS FOR TOOL INPUT/OUTPUT
# =============================================================================

class FileInfo(BaseModel):
    """Information about a generated file"""
    path: str = Field(description="Full path to the file")
    filename: str = Field(description="Just the filename with extension")
    type: str = Field(description="Type of file: new_rules_delta, comparison_audit, or comparison_audit_formatted")
    purpose: str = Field(description="Whether this is PRIMARY OUTPUT or AUDIT ONLY")
    row_count: int = Field(description="Number of rows in the file")
    preview: str = Field(description="Text preview of first 10 rows")


class ComparisonSummary(BaseModel):
    """Statistics about the comparison results"""
    total_proposed_rules: int = Field(description="Number of rules in proposed contract")
    new_rules_identified: int = Field(description="Number of new rules that need to be added (the delta)")
    rules_already_covered: int = Field(description="Number of proposed rules already covered by master")
    conflicts_detected: int = Field(description="Number of rules that conflict with master")
    rules_needing_review: int = Field(description="Number of rules flagged for human review (confidence < 0.9)")
    business_terms_analyzed: int = Field(description="Total business terms analyzed")
    existing_terms_compared: int = Field(description="Business terms that exist in both contracts")
    new_terms_consolidated: int = Field(description="Business terms only in proposed contract (new)")


class CompareContractsOutput(BaseModel):
    """Output from the compare_contracts tool"""
    status: str = Field(description="'success' or 'error'")
    message: str = Field(description="Human-readable summary of what happened")
    files_created: Optional[List[FileInfo]] = Field(
        default=None,
        description="List of files created. First file is always new_rules_delta (PRIMARY OUTPUT). "
                    "Subsequent files are audit trails (AUDIT ONLY)."
    )
    summary: Optional[ComparisonSummary] = Field(
        default=None,
        description="Detailed statistics about comparison results"
    )
    warnings: List[str] = Field(
        default_factory=list,
        description="Non-fatal warnings encountered during processing"
    )
    errors: List[str] = Field(
        default_factory=list,
        description="Partial errors (processing continued despite these)"
    )
    error_type: Optional[str] = Field(
        default=None,
        description="Type of error if status is 'error'"
    )
    suggestion: Optional[str] = Field(
        default=None,
        description="Suggestion for resolving the error"
    )


# =============================================================================
# UTILITY: FLEXIBLE COLUMN MATCHING (same as Tool 1)
# =============================================================================

def find_column_name(df: pd.DataFrame, target_column: str) -> Optional[str]:
    """Find the actual column name in DataFrame that matches our target"""
    if target_column not in COLUMN_ALIASES:
        return target_column if target_column in df.columns else None
    
    possible_names = COLUMN_ALIASES[target_column]
    df_columns_lower = {col.lower().strip(): col for col in df.columns}
    
    for possible_name in possible_names:
        possible_lower = possible_name.lower().strip()
        if possible_lower in df_columns_lower:
            return df_columns_lower[possible_lower]
    
    return None


def standardize_dataframe_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Rename DataFrame columns to standard names"""
    rename_map = {}
    
    for standard_name in ['business_rule', 'business_term', 'description']:
        actual_name = find_column_name(df, standard_name)
        if actual_name and actual_name != standard_name:
            rename_map[actual_name] = standard_name
    
    if rename_map:
        df = df.rename(columns=rename_map)
    
    return df


# =============================================================================
# SYSTEM PROMPTS
# =============================================================================

def get_comparison_system_prompt():
    """System prompt for comparing proposed rules against master contract rules"""
    return """You are a data quality expert analyzing business rules to compare proposed contract rules against existing master contract rules. Your goal is to identify which proposed rules are genuinely new versus already covered by the master contract.

## INPUT FORMAT

You will receive:
1. **Master Contract Rules**: Existing rules from the producer's golden contract
2. **Proposed Contract Rules**: New rules requested by a consumer

All rules are formatted as "[Business Term]: [Rule Description]"

**CRITICAL**: If a rule starts with text that matches or is similar to the business term, ignore that prefix. Only analyze the actual rule logic after the colon.

## COMPARISON LOGIC

For each proposed rule, determine its relationship to the master contract rules:

### 1. DUPLICATE WITHIN PROPOSED
First, identify if the proposed rule is a duplicate of another proposed rule.

### 2. ALREADY COVERED
The proposed rule is already covered if ANY master rule enforces the same or stricter constraint.

### 3. NEW RULES NEEDED
The proposed rule requires addition if it represents genuinely new business logic:
- **new_stricter**: Proposed rule is stricter than any master rule
- **new_different**: Different constraint type not covered by master

### 4. CONFLICTS
Flag potential conflicts where proposed and master rules contradict each other.

## CONFIDENCE SCORING

- **1.0**: Absolutely certain
- **0.9-0.99**: Very confident
- **0.8-0.89**: Confident with slight ambiguity
- **< 0.8**: Uncertain - flag for human review

## OUTPUT FORMAT

Return ONLY valid JSON with no markdown formatting:

{
    "business_term": "string",
    "master_rules_count": 0,
    "proposed_rules_count": 0,
    "new_rules_required": false,
    "analysis": [
        {
            "proposed_rule": "exact proposed rule text",
            "proposed_rule_index": 0,
            "status": "covered|new_stricter|new_different|conflict|duplicate_within_proposed",
            "matched_master_rule": "master rule text or null",
            "matched_master_index": 0,
            "canonical_rule": "rule to add if new, or master rule if covered",
            "confidence": 0.95,
            "reasoning": "clear explanation of decision"
        }
    ],
    "notes": "any additional observations"
}"""


def get_consolidation_system_prompt():
    """System prompt for consolidating rules within new business terms"""
    return """You are a data quality expert analyzing business rules to identify and eliminate true duplicates. Your goal is to ensure each unique business logic constraint is preserved while removing redundant rules.

## WHAT MAKES RULES DUPLICATES

Two rules are duplicates ONLY if they enforce the exact same business logic constraint.

## OUTPUT FORMAT

Return ONLY valid JSON with no markdown formatting:

{
    "business_term": "string",
    "total_rules_analyzed": 0,
    "unique_rules_identified": 0,
    "duplicates_found": false,
    "analysis": [
        {
            "original_rule": "exact input text",
            "original_rule_index": 0,
            "status": "unique|duplicate",
            "canonical_rule": "best version of rule",
            "confidence": 0.95,
            "consolidated_with": ["array of other rule texts if duplicate"],
            "duplicate_of_index": null,
            "reasoning": "clear explanation of decision"
        }
    ],
    "notes": "any additional observations"
}"""


# =============================================================================
# DATA LOADING AND VALIDATION
# =============================================================================

def load_and_validate_contract(csv_path: str, contract_type: str) -> tuple[pd.DataFrame, List[str]]:
    """
    Load and validate a contract CSV file
    
    Parameters:
        csv_path: Path to CSV file
        contract_type: "master" or "proposed" (for error messages)
        
    Returns:
        Tuple of (cleaned_df, warnings_list)
    """
    warnings = []
    
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"{contract_type.capitalize()} contract file not found: {csv_path}")
    
    df = pd.read_csv(csv_path)
    original_count = len(df)
    
    # Standardize column names
    df = standardize_dataframe_columns(df)
    
    # Check required columns
    required_columns = ['business_rule', 'business_term']
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"{contract_type.capitalize()} contract missing required columns: {missing_columns}. "
                        f"Available columns: {list(df.columns)}")
    
    # Remove rows with null values
    df = df.dropna(subset=['business_rule', 'business_term'])
    null_removed = original_count - len(df)
    if null_removed > 0:
        warnings.append(f"{contract_type.capitalize()}: Removed {null_removed} rows with null values")
    
    # Trim whitespace
    df['business_rule'] = df['business_rule'].str.strip()
    df['business_term'] = df['business_term'].str.strip()
    
    # Remove exact duplicates
    before_dedup = len(df)
    df = df.drop_duplicates(subset=['business_rule', 'business_term'])
    duplicates_removed = before_dedup - len(df)
    if duplicates_removed > 0:
        warnings.append(f"{contract_type.capitalize()}: Removed {duplicates_removed} exact duplicate rows")
    
    if len(df) == 0:
        raise ValueError(f"No valid data remaining in {contract_type} contract after cleaning")
    
    return df, warnings


# =============================================================================
# BUSINESS TERM CATEGORIZATION
# =============================================================================

def categorize_business_terms(master_df: pd.DataFrame, proposed_df: pd.DataFrame) -> Dict[str, List[str]]:
    """
    Categorize business terms into existing (in both) and new (only in proposed)
    
    Parameters:
        master_df: Master contract DataFrame
        proposed_df: Proposed contract DataFrame
        
    Returns:
        Dict with 'existing_terms' and 'new_terms' lists
    """
    master_terms = set(master_df['business_term'].unique())
    proposed_terms = set(proposed_df['business_term'].unique())
    
    existing_terms = list(master_terms.intersection(proposed_terms))
    existing_terms.sort()
    
    new_terms = list(proposed_terms - master_terms)
    new_terms.sort()
    
    return {
        "existing_terms": existing_terms,
        "new_terms": new_terms
    }


def get_rules_for_business_term(df: pd.DataFrame, business_term: str) -> pd.DataFrame:
    """Filter dataframe to only rules for a specific business term"""
    return df[df['business_term'] == business_term].copy()


# =============================================================================
# LLM INVOCATION
# =============================================================================

def invoke_bedrock_model(system_prompt: str, user_prompt: str) -> str:
    """Call AWS Bedrock API to invoke Claude"""
    bedrock = boto3.client(
        service_name='bedrock-runtime',
        region_name='us-east-1'
    )
    
    body = {
        "modelId": "global.anthropic.claude-haiku-4-5-20251001-v1:0",
        "inferenceConfig": {
            "maxTokens": 4096,
            "temperature": 0
        },
        "system": [{"text": system_prompt}],
        "messages": [{
            'role': 'user',
            'content': [{'text': user_prompt}]
        }]
    }
    
    response = bedrock.converse(**body)
    result = response['output']['message']['content'][0]['text']
    
    return result


def validate_comparison_response(response: Dict, business_term: str):
    """Validate comparison LLM response structure"""
    if 'business_term' not in response:
        raise ValueError("Missing 'business_term' field")
    
    if 'analysis' not in response:
        raise ValueError("Missing 'analysis' field")
    
    if not isinstance(response['analysis'], list):
        raise ValueError("'analysis' must be a list")
    
    for i, entry in enumerate(response['analysis']):
        required_fields = ['proposed_rule', 'status', 'canonical_rule', 'confidence']
        for field in required_fields:
            if field not in entry:
                raise ValueError(f"Analysis entry {i} missing field: {field}")
        
        valid_statuses = ['covered', 'new_stricter', 'new_different', 'conflict', 'duplicate_within_proposed']
        if entry['status'] not in valid_statuses:
            raise ValueError(f"Analysis entry {i} has invalid status: {entry['status']}")
        
        if not (0.0 <= entry['confidence'] <= 1.0):
            raise ValueError(f"Analysis entry {i} has invalid confidence: {entry['confidence']}")


def validate_consolidation_response(response: Dict, business_term: str):
    """Validate consolidation LLM response structure"""
    if 'business_term' not in response:
        raise ValueError("Missing 'business_term' field")
    
    if 'analysis' not in response:
        raise ValueError("Missing 'analysis' field")
    
    if not isinstance(response['analysis'], list):
        raise ValueError("'analysis' must be a list")
    
    for i, entry in enumerate(response['analysis']):
        required_fields = ['original_rule', 'status', 'canonical_rule', 'confidence']
        for field in required_fields:
            if field not in entry:
                raise ValueError(f"Analysis entry {i} missing field: {field}")
        
        if entry['status'] not in ['unique', 'duplicate']:
            raise ValueError(f"Analysis entry {i} has invalid status: {entry['status']}")
        
        if not (0.0 <= entry['confidence'] <= 1.0):
            raise ValueError(f"Analysis entry {i} has invalid confidence: {entry['confidence']}")


# =============================================================================
# COMPARISON LOGIC - EXISTING TERMS
# =============================================================================

def compare_rules_for_existing_term(master_rules_df: pd.DataFrame,
                                    proposed_rules_df: pd.DataFrame,
                                    business_term: str) -> Dict[str, Any]:
    """
    Compare proposed rules against master rules for an existing business term
    
    Parameters:
        master_rules_df: All master rules for this business term
        proposed_rules_df: All proposed rules for this business term
        business_term: The business term being compared
        
    Returns:
        Dictionary with structured JSON comparison analysis
    """
    master_rules = master_rules_df['business_rule'].tolist()
    proposed_rules = proposed_rules_df['business_rule'].tolist()
    
    user_prompt = f"""business_term: {business_term}

Master Contract Rules:
{json.dumps(master_rules, indent=2)}

Proposed Contract Rules:
{json.dumps(proposed_rules, indent=2)}

Compare each proposed rule against the master rules. Return your analysis as valid JSON only (no markdown, no code blocks)."""
    
    system_prompt = get_comparison_system_prompt()
    
    for attempt in range(MAX_RETRIES):
        try:
            llm_response = invoke_bedrock_model(system_prompt, user_prompt)
            
            # Clean markdown if present
            llm_response = llm_response.strip()
            if llm_response.startswith("```"):
                lines = llm_response.split('\n')
                llm_response = '\n'.join(lines[1:-1])
            
            result = json.loads(llm_response)
            validate_comparison_response(result, business_term)
            
            return result
            
        except (json.JSONDecodeError, ValueError, KeyError) as e:
            if attempt < MAX_RETRIES - 1:
                user_prompt += "\n\nIMPORTANT: Return ONLY valid JSON. No markdown code blocks, no explanations outside JSON."
            else:
                return {
                    "business_term": business_term,
                    "error": f"Failed to parse LLM response after {MAX_RETRIES} attempts: {str(e)}",
                    "master_rules_count": len(master_rules),
                    "proposed_rules_count": len(proposed_rules),
                    "new_rules_required": False,
                    "analysis": []
                }


# =============================================================================
# CONSOLIDATION LOGIC - NEW TERMS
# =============================================================================

def consolidate_rules_for_new_term(proposed_rules_df: pd.DataFrame, business_term: str) -> Dict[str, Any]:
    """
    Consolidate proposed rules for a NEW business term (not in master)
    
    Parameters:
        proposed_rules_df: All proposed rules for this new term
        business_term: The business term
        
    Returns:
        Dictionary with structured JSON consolidation analysis
    """
    rules_list = proposed_rules_df['business_rule'].tolist()
    
    # If only 1 rule, no LLM call needed
    if len(rules_list) == 1:
        return {
            "business_term": business_term,
            "total_rules_analyzed": 1,
            "unique_rules_identified": 1,
            "duplicates_found": False,
            "analysis": [
                {
                    "original_rule": rules_list[0],
                    "original_rule_index": 0,
                    "status": "unique",
                    "canonical_rule": rules_list[0],
                    "confidence": 1.0,
                    "consolidated_with": [],
                    "duplicate_of_index": None,
                    "reasoning": "Only one rule for this new business term. No consolidation needed."
                }
            ],
            "notes": "Single rule for new business term - no LLM processing required"
        }
    
    # Multiple rules - use LLM
    user_prompt = f"""Business Term: {business_term}

Rules to analyze:
{json.dumps(rules_list, indent=2)}

Analyze these rules and identify semantic duplicates. Return your analysis as valid JSON only (no markdown, no code blocks)."""
    
    system_prompt = get_consolidation_system_prompt()
    
    for attempt in range(MAX_RETRIES):
        try:
            llm_response = invoke_bedrock_model(system_prompt, user_prompt)
            
            # Clean markdown if present
            llm_response = llm_response.strip()
            if llm_response.startswith("```"):
                lines = llm_response.split('\n')
                llm_response = '\n'.join(lines[1:-1])
            
            result = json.loads(llm_response)
            validate_consolidation_response(result, business_term)
            
            return result
            
        except (json.JSONDecodeError, ValueError, KeyError) as e:
            if attempt < MAX_RETRIES - 1:
                user_prompt += "\n\nIMPORTANT: Return ONLY valid JSON. No markdown code blocks, no explanations outside JSON."
            else:
                return {
                    "business_term": business_term,
                    "error": f"Failed to parse LLM response after {MAX_RETRIES} attempts: {str(e)}",
                    "total_rules_analyzed": len(rules_list),
                    "unique_rules_identified": 0,
                    "duplicates_found": False,
                    "analysis": []
                }


# =============================================================================
# ORCHESTRATION
# =============================================================================

def process_all_business_terms(master_df: pd.DataFrame,
                               proposed_df: pd.DataFrame,
                               categorized_terms: Dict[str, List[str]]) -> List[Dict[str, Any]]:
    """
    Process all business terms - both existing (comparison) and new (consolidation)
    
    Parameters:
        master_df: Cleaned master contract
        proposed_df: Cleaned proposed contract
        categorized_terms: Dict with 'existing_terms' and 'new_terms' lists
        
    Returns:
        List of analysis results for all terms
    """
    all_results = []
    
    # Process existing terms (comparison against master)
    existing_terms = categorized_terms["existing_terms"]
    for business_term in existing_terms:
        master_rules_df = get_rules_for_business_term(master_df, business_term)
        proposed_rules_df = get_rules_for_business_term(proposed_df, business_term)
        
        result = compare_rules_for_existing_term(master_rules_df, proposed_rules_df, business_term)
        result['term_type'] = 'existing'
        all_results.append(result)
    
    # Process new terms (consolidation within proposed)
    new_terms = categorized_terms["new_terms"]
    for business_term in new_terms:
        proposed_rules_df = get_rules_for_business_term(proposed_df, business_term)
        
        result = consolidate_rules_for_new_term(proposed_rules_df, business_term)
        result['term_type'] = 'new'
        all_results.append(result)
    
    return all_results


# =============================================================================
# DELTA OUTPUT CREATION
# =============================================================================

def create_delta_outputs(all_results: List[Dict[str, Any]]) -> tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    """
    Create delta outputs from analysis results
    
    Parameters:
        all_results: Analysis results from all business terms
        
    Returns:
        Tuple of (new_rules_df, audit_df, errors_list)
    """
    new_rules = []
    audit_records = []
    errors = []
    
    for result in all_results:
        business_term = result['business_term']
        term_type = result.get('term_type', 'unknown')
        
        # Track errors but continue
        if 'error' in result:
            errors.append(f"{business_term}: {result['error']}")
            continue
        
        seen_canonical = set()
        
        for rule_analysis in result['analysis']:
            # Create audit record
            audit_record = {
                'business_term': business_term,
                'term_type': term_type,
                'original_rule': rule_analysis.get('original_rule', rule_analysis.get('proposed_rule', '')),
                'status': rule_analysis['status'],
                'canonical_rule': rule_analysis['canonical_rule'],
                'confidence': rule_analysis['confidence'],
                'needs_review': rule_analysis['confidence'] < 0.9,
                'reasoning': rule_analysis['reasoning']
            }
            
            # Add comparison-specific fields for existing terms
            if term_type == 'existing':
                audit_record['matched_master_rule'] = rule_analysis.get('matched_master_rule')
                audit_record['matched_master_index'] = rule_analysis.get('matched_master_index')
            
            audit_records.append(audit_record)
            
            # Add to new rules if needed (INCLUDING CONFLICTS)
            canonical_rule = rule_analysis['canonical_rule']
            status = rule_analysis['status']
            
            # Rules that need to be added to master contract
            if status in ['new_stricter', 'new_different', 'conflict'] or (term_type == 'new' and status == 'unique'):
                if canonical_rule not in seen_canonical:
                    seen_canonical.add(canonical_rule)
                    
                    original_rule_count = sum(
                        1 for r in result['analysis']
                        if r['canonical_rule'] == canonical_rule
                    )
                    
                    new_rules.append({
                        'business_rule': canonical_rule,
                        'business_term': business_term,
                        'status': status,
                        'term_type': term_type,
                        'related_master_rule': rule_analysis.get('matched_master_rule'),
                        'related_master_index': rule_analysis.get('matched_master_index'),
                        'rule_source': 'comparison_analysis',
                        'original_rule_count': original_rule_count,
                        'confidence': rule_analysis['confidence'],
                        'needs_review': rule_analysis['confidence'] < 0.9,
                        'reasoning': rule_analysis['reasoning']
                    })
    
    new_rules_df = pd.DataFrame(new_rules)
    audit_df = pd.DataFrame(audit_records)
    
    # Sort for consistency
    if not new_rules_df.empty:
        new_rules_df = new_rules_df.sort_values(['business_term', 'business_rule']).reset_index(drop=True)
    
    if not audit_df.empty:
        audit_df = audit_df.sort_values(
            ['business_term', 'needs_review', 'original_rule'],
            ascending=[True, False, True]
        ).reset_index(drop=True)
    
    return new_rules_df, audit_df, errors


# =============================================================================
# FILE PREVIEW
# =============================================================================

def generate_file_preview(df: pd.DataFrame, max_rows: int = 10) -> str:
    """Generate a text preview of the first N rows of a dataframe"""
    if df.empty:
        return "Empty DataFrame - no rules to add"
    preview_df = df.head(max_rows)
    return preview_df.to_string(index=False)


# =============================================================================
# OUTPUT GENERATION
# =============================================================================

def save_outputs(new_rules_df: pd.DataFrame,
                audit_df: pd.DataFrame,
                output_dir: str,
                timestamp: str) -> List[Dict[str, Any]]:
    """
    Save delta and audit outputs
    
    Parameters:
        new_rules_df: New rules that need to be added (the delta)
        audit_df: Detailed comparison audit trail
        output_dir: Directory to save outputs
        timestamp: Timestamp string for unique filenames
        
    Returns:
        List of file information dicts
    """
    os.makedirs(output_dir, exist_ok=True)
    
    files_created = []
    
    # Save new rules delta (PRIMARY OUTPUT)
    delta_filename = f'new_rules_delta_{timestamp}.csv'
    delta_path = os.path.join(output_dir, delta_filename)
    new_rules_df.to_csv(delta_path, index=False)
    
    files_created.append({
        'path': delta_path,
        'filename': delta_filename,
        'type': 'new_rules_delta',
        'purpose': 'PRIMARY OUTPUT - Use this as input to merge_and_highlight function',
        'row_count': len(new_rules_df),
        'preview': generate_file_preview(new_rules_df, max_rows=10)
    })
    
    # Save comparison audit (AUDIT ONLY)
    audit_filename = f'comparison_audit_{timestamp}.csv'
    audit_path = os.path.join(output_dir, audit_filename)
    audit_df.to_csv(audit_path, index=False)
    
    files_created.append({
        'path': audit_path,
        'filename': audit_filename,
        'type': 'comparison_audit',
        'purpose': 'AUDIT ONLY - Do NOT use as input to other functions. For human review only.',
        'row_count': len(audit_df),
        'preview': generate_file_preview(audit_df, max_rows=10)
    })
    
    # Try to save formatted Excel (AUDIT ONLY)
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        
        excel_filename = f'comparison_audit_formatted_{timestamp}.xlsx'
        excel_path = os.path.join(output_dir, excel_filename)
        audit_df.to_excel(excel_path, index=False, engine='openpyxl')
        
        # Apply formatting
        wb = load_workbook(excel_path)
        ws = wb.active
        
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        needs_review_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'needs_review':
                needs_review_col = idx
                break
        
        if needs_review_col:
            for row in range(2, ws.max_row + 1):
                if ws.cell(row, needs_review_col).value == True:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row, col).fill = yellow_fill
        
        wb.save(excel_path)
        
        files_created.append({
            'path': excel_path,
            'filename': excel_filename,
            'type': 'comparison_audit_formatted',
            'purpose': 'AUDIT ONLY - Formatted Excel with highlighting. Do NOT use as input to other functions.',
            'row_count': len(audit_df),
            'preview': 'Excel file with yellow highlighting for low-confidence comparisons'
        })
        
    except ImportError:
        pass  # openpyxl not available, skip Excel
    
    return files_created


# =============================================================================
# MAIN TOOL FUNCTION - LANGGRAPH @tool DECORATOR
# =============================================================================

@tool
def compare_contracts(
    master_contract_path: str,
    proposed_contract_path: str,
    output_dir: str = "./output"
) -> dict:
    """
    Compare proposed contract against master contract to identify new rules (delta detection).
    
    Performs semantic comparison between a proposed consumer contract and the master/golden 
    contract to identify which rules are genuinely new versus already covered. Uses LLM 
    analysis to detect semantic equivalence, stricter requirements, and conflicts.
    
    IMPORTANT:
    - The new_rules_delta file is the PRIMARY OUTPUT - use this as input to merge_and_highlight
    - Audit files are for HUMAN REVIEW ONLY - do not pass to other functions
    - Conflicts are included in the delta and flagged for human review
    - All files are timestamped for uniqueness
    
    Args:
        master_contract_path: Full path to master/golden contract CSV file. Must contain columns:
                            business_rule and business_term (case-insensitive).
        proposed_contract_path: Full path to proposed consumer contract CSV file. Must contain 
                              columns: business_rule and business_term (case-insensitive).
        output_dir: Directory where output files will be saved. Defaults to './output'
    
    Returns:
        Dictionary with:
        - status: 'success' or 'error'
        - message: Human-readable summary
        - files_created: List of file objects (first is always PRIMARY OUTPUT new_rules_delta)
          Each file has: path (full path), filename, type, purpose, row_count, preview
        - summary: Statistics dict with total_proposed_rules, new_rules_identified, 
          rules_already_covered, conflicts_detected, rules_needing_review, business_terms_analyzed
        - warnings: List of non-fatal warnings
        - errors: List of partial errors (if any business terms failed to process)
        - error_type: (on error) Type of error encountered
        - suggestion: (on error) Suggestion for resolving the error
    """
    
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        all_warnings = []
        
        # Step 1: Load and validate both contracts
        master_df, master_warnings = load_and_validate_contract(master_contract_path, "master")
        all_warnings.extend(master_warnings)
        
        proposed_df, proposed_warnings = load_and_validate_contract(proposed_contract_path, "proposed")
        all_warnings.extend(proposed_warnings)
        
        # Step 2: Categorize business terms
        categorized_terms = categorize_business_terms(master_df, proposed_df)
        
        # Step 3: Process all business terms
        all_results = process_all_business_terms(master_df, proposed_df, categorized_terms)
        
        # Step 4: Create delta outputs
        new_rules_df, audit_df, processing_errors = create_delta_outputs(all_results)
        
        # Step 5: Save outputs
        files_created = save_outputs(new_rules_df, audit_df, output_dir, timestamp)
        
        # Convert file dicts to FileInfo objects then back to dicts
        file_info_objects = [FileInfo(**file_dict) for file_dict in files_created]
        
        # Generate summary
        conflicts_count = len(new_rules_df[new_rules_df['status'] == 'conflict']) if not new_rules_df.empty else 0
        
        summary = ComparisonSummary(
            total_proposed_rules=len(proposed_df),
            new_rules_identified=len(new_rules_df),
            rules_already_covered=len(audit_df[audit_df['status'] == 'covered']) if not audit_df.empty else 0,
            conflicts_detected=conflicts_count,
            rules_needing_review=len(audit_df[audit_df['needs_review'] == True]) if not audit_df.empty else 0,
            business_terms_analyzed=len(audit_df['business_term'].unique()) if not audit_df.empty else 0,
            existing_terms_compared=len(categorized_terms['existing_terms']),
            new_terms_consolidated=len(categorized_terms['new_terms'])
        )
        
        # Construct success message
        if len(new_rules_df) == 0:
            message = (
                f"No new rules needed! All {summary.total_proposed_rules} proposed rules are "
                f"already covered by the master contract."
            )
        else:
            message = (
                f"Comparison complete. Identified {summary.new_rules_identified} new rules "
                f"(delta) out of {summary.total_proposed_rules} proposed rules. "
                f"{summary.rules_already_covered} rules already covered by master contract."
            )
            
            if conflicts_count > 0:
                message += f" WARNING: {conflicts_count} conflicts detected requiring human review."
            
            if summary.rules_needing_review > 0:
                message += f" {summary.rules_needing_review} rules flagged for review (confidence < 0.9)."
        
        output = CompareContractsOutput(
            status='success',
            message=message,
            files_created=file_info_objects,
            summary=summary,
            warnings=all_warnings,
            errors=processing_errors if processing_errors else []
        )
        return output.model_dump()
        
    except FileNotFoundError as e:
        output = CompareContractsOutput(
            status='error',
            message=f"File not found: {str(e)}",
            error_type='FileNotFoundError',
            suggestion='Verify both contract file paths are correct and files exist.'
        )
        return output.model_dump()
        
    except ValueError as e:
        output = CompareContractsOutput(
            status='error',
            message=f"Validation error: {str(e)}",
            error_type='ValueError',
            suggestion='Check that both CSV files have required columns (business_rule, business_term) and contain valid data.'
        )
        return output.model_dump()
        
    except Exception as e:
        output = CompareContractsOutput(
            status='error',
            message=f"Unexpected error during comparison: {str(e)}",
            error_type=type(e).__name__,
            suggestion='Review the error message and input data. Contact support if the issue persists.'
        )
        return output.model_dump()


# =============================================================================
# END OF TOOL - Schema auto-generated by @tool decorator from Pydantic models
# =============================================================================


# =============================================================================
# USAGE EXAMPLES
# =============================================================================

"""
HOW TO USE THIS TOOL IN LANGGRAPH:

1. Import the tool:
   from compare_contracts_tool import compare_contracts

2. Add to your agent's tools list:
   tools = [consolidate_contract, compare_contracts, merge_and_highlight]

3. The agent can invoke it (typically after consolidate_contract):
   # First, get master contract from Tool 1
   result1 = consolidate_contract.invoke({"input_csv_path": "consumer_data.csv"})
   master_path = result1['files_created'][0]['path']
   
   # Then compare proposed contract against master
   result2 = compare_contracts.invoke({
       "master_contract_path": master_path,
       "proposed_contract_path": "proposed_contract.csv"
   })

4. Access structured results (returns dict):
   if result2['status'] == "success":
       # First file is PRIMARY OUTPUT (new rules delta)
       delta_path = result2['files_created'][0]['path']
       print(f"Delta file: {delta_path}")
       print(f"New rules to add: {result2['summary']['new_rules_identified']}")
       
       # Check for conflicts
       if result2['summary']['conflicts_detected'] > 0:
           print(f"WARNING: {result2['summary']['conflicts_detected']} conflicts need review!")
       
       # Audit files are for human review only
       for file_info in result2['files_created'][1:]:
           if "AUDIT ONLY" in file_info['purpose']:
               print(f"Audit: {file_info['path']} (do not pass to other tools)")

IMPORTANT NOTES FOR AGENT BEHAVIOR:
- Only use the new_rules_delta file (first in files_created) as input to merge_and_highlight
- Audit files are marked with purpose="AUDIT ONLY" - never pass these to other tools
- Conflicts are INCLUDED in the delta file and marked with status='conflict'
- If new_rules_identified is 0, no merge is needed - proposed contract fully covered
- If status == "error", check error_type and suggestion for recovery steps
- All file paths are FULL PATHS ready to use in subsequent tool calls
"""


# =============================================================================
# FOR STANDALONE TESTING (NOT PART OF TOOL)
# =============================================================================

if __name__ == "__main__":
    # Test the tool standalone
    result = compare_contracts.invoke({
        "master_contract_path": "output/master_contract_20250524.csv",
        "proposed_contract_path": "test_proposed_contract.csv"
    })
    
    print(f"\nStatus: {result['status']}")
    print(f"Message: {result['message']}")
    
    if result['status'] == "success":
        print(f"\nFiles created:")
        for file_info in result['files_created']:
            print(f"  - {file_info['filename']} ({file_info['type']})")
            print(f"    Purpose: {file_info['purpose']}")
            print(f"    Rows: {file_info['row_count']}")
        
        print(f"\nSummary:")
        print(f"  Proposed rules: {result['summary']['total_proposed_rules']}")
        print(f"  New rules (delta): {result['summary']['new_rules_identified']}")
        print(f"  Already covered: {result['summary']['rules_already_covered']}")
        print(f"  Conflicts: {result['summary']['conflicts_detected']}")
    else:
        print(f"\nError: {result['error_type']}")
        print(f"Suggestion: {result['suggestion']}")