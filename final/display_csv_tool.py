"""
LANGGRAPH UTILITY TOOL: Display CSV
Display CSV/Excel files in the UI artifact panel with highlighting support
"""

import pandas as pd
import os
from typing import Dict, List, Any, Optional
from datetime import datetime
from pydantic import BaseModel, Field
from langchain_core.tools import tool

# Try to import openpyxl for Excel highlighting extraction
try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# =============================================================================
# CONFIGURATION
# =============================================================================

# Colors from merge_and_highlight tool
# These are the actual colors used, checking multiple variants for flexibility
CONFLICT_RED_VARIANTS = ["FF0000", "FFCDD2", "FFCCCC", "FF9999", "EF9A9A"]
NEW_RULE_YELLOW_VARIANTS = ["FFFF00", "FFFF99", "FFFF66", "FFFFCC", "FFF59D"]


# =============================================================================
# PYDANTIC MODELS
# =============================================================================

class HighlightInfo(BaseModel):
    """Information about row highlighting"""
    has_highlighting: bool = Field(description="Whether the file has highlighted rows")
    conflict_rows: List[int] = Field(default_factory=list, description="Row indices with conflict highlighting (red)")
    new_rule_rows: List[int] = Field(default_factory=list, description="Row indices with new rule highlighting (yellow)")


class DisplayMetadata(BaseModel):
    """Metadata about the displayed file"""
    total_rows: int = Field(description="Total number of rows")
    total_columns: int = Field(description="Total number of columns")
    columns: List[str] = Field(description="Column names")
    file_size_mb: float = Field(description="File size in MB")
    file_type: str = Field(description="File type (csv or excel)")


class DisplayCsvOutput(BaseModel):
    """Output from the display_csv tool - designed for UI artifact interception"""
    display_type: str = Field(default="dataframe", description="Display type for UI interception")
    status: str = Field(description="'success' or 'error'")
    message: str = Field(description="Human-readable summary")
    title: str = Field(description="Display title for the artifact")
    file_path: str = Field(description="Full path to the source file")
    html_content: Optional[str] = Field(default=None, description="HTML table with styling for display")
    csv_data: Optional[str] = Field(default=None, description="Raw CSV data for download")
    metadata: Optional[DisplayMetadata] = Field(default=None, description="File metadata")
    highlight_info: Optional[HighlightInfo] = Field(default=None, description="Highlighting information")
    timestamp: str = Field(description="When the artifact was created")
    error_type: Optional[str] = Field(default=None, description="Error type if failed")
    suggestion: Optional[str] = Field(default=None, description="Suggestion for fixing error")


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def get_file_size_mb(file_path: str) -> float:
    """Get file size in megabytes"""
    try:
        size_bytes = os.path.getsize(file_path)
        return round(size_bytes / (1024 * 1024), 2)
    except:
        return 0.0


def get_file_type(file_path: str) -> str:
    """Determine file type from extension"""
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.csv':
        return 'csv'
    elif ext in ['.xlsx', '.xls', '.xlsm']:
        return 'excel'
    else:
        raise ValueError(f"Unsupported file type: {ext}")


def extract_excel_highlighting(file_path: str) -> Dict[str, List[int]]:
    """
    Extract row highlighting information from Excel file.
    
    Returns:
        Dict with 'conflict_rows' and 'new_rule_rows' lists
    """
    if not OPENPYXL_AVAILABLE:
        return {'conflict_rows': [], 'new_rule_rows': []}
    
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        conflict_rows = []
        new_rule_rows = []
        
        # Iterate through rows (skip header, row 1 is header in Excel)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=0):
            # Check first cell's fill color
            first_cell = row[0]
            
            if first_cell.fill and first_cell.fill.fill_type == "solid":
                fg_color = first_cell.fill.fgColor
                
                if fg_color and fg_color.rgb:
                    color = fg_color.rgb
                    
                    # Handle different color formats
                    if isinstance(color, str):
                        # Remove alpha channel if present (ARGB -> RGB)
                        # Also handle "00XXXXXX" format where first two chars are alpha
                        if len(color) == 8:
                            color = color[2:]
                        
                        color_upper = color.upper()
                        
                        # Check for red variants (conflict)
                        if color_upper in CONFLICT_RED_VARIANTS:
                            conflict_rows.append(row_idx)
                        # Check for yellow variants (new rule)
                        elif color_upper in NEW_RULE_YELLOW_VARIANTS:
                            new_rule_rows.append(row_idx)
        
        return {
            'conflict_rows': conflict_rows,
            'new_rule_rows': new_rule_rows
        }
    
    except Exception as e:
        print(f"Warning: Could not extract highlighting: {e}")
        import traceback
        traceback.print_exc()
        return {'conflict_rows': [], 'new_rule_rows': []}


def dataframe_to_styled_html(
    df: pd.DataFrame,
    conflict_rows: List[int] = None,
    new_rule_rows: List[int] = None,
    title: str = None
) -> str:
    """
    Convert DataFrame to HTML table with row highlighting.
    
    Args:
        df: DataFrame to convert
        conflict_rows: Row indices to highlight red (conflicts)
        new_rule_rows: Row indices to highlight yellow (new rules)
        title: Optional title for the table
        
    Returns:
        HTML string with styled table
    """
    conflict_rows = conflict_rows or []
    new_rule_rows = new_rule_rows or []
    
    # Build HTML
    html_parts = []
    
    # Add custom styles
    html_parts.append("""
    <style>
        .artifact-table-container {
            max-height: 600px;
            overflow-y: auto;
            border: 1px solid #e5e7eb;
            border-radius: 8px;
        }
        .artifact-table {
            width: 100%;
            border-collapse: collapse;
            font-size: 13px;
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
        }
        .artifact-table th {
            background-color: #f3f4f6;
            padding: 10px 12px;
            text-align: left;
            font-weight: 600;
            border-bottom: 2px solid #e5e7eb;
            position: sticky;
            top: 0;
            z-index: 10;
        }
        .artifact-table td {
            padding: 8px 12px;
            border-bottom: 1px solid #e5e7eb;
            max-width: 300px;
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
        }
        .artifact-table tr:hover {
            background-color: #f9fafb;
        }
        .row-conflict {
            background-color: #FF0000 !important;
            color: white;
        }
        .row-conflict:hover {
            background-color: #CC0000 !important;
        }
        .row-new-rule {
            background-color: #FFFF00 !important;
        }
        .row-new-rule:hover {
            background-color: #E6E600 !important;
        }
        .highlight-legend {
            display: flex;
            gap: 20px;
            margin-bottom: 10px;
            font-size: 12px;
        }
        .legend-item {
            display: flex;
            align-items: center;
            gap: 6px;
        }
        .legend-color {
            width: 16px;
            height: 16px;
            border-radius: 3px;
            border: 1px solid #ccc;
        }
        .legend-red { background-color: #FF0000; }
        .legend-yellow { background-color: #FFFF00; }
    </style>
    """)
    
    # Add legend if there are highlighted rows
    if conflict_rows or new_rule_rows:
        html_parts.append('<div class="highlight-legend">')
        if conflict_rows:
            html_parts.append(f'''
                <div class="legend-item">
                    <div class="legend-color legend-red"></div>
                    <span>🚨 Conflicts ({len(conflict_rows)} rows) - Requires immediate review</span>
                </div>
            ''')
        if new_rule_rows:
            html_parts.append(f'''
                <div class="legend-item">
                    <div class="legend-color legend-yellow"></div>
                    <span>📝 New Rules ({len(new_rule_rows)} rows) - Needs approval</span>
                </div>
            ''')
        html_parts.append('</div>')
    
    # Start table
    html_parts.append('<div class="artifact-table-container">')
    html_parts.append('<table class="artifact-table">')
    
    # Header row
    html_parts.append('<thead><tr>')
    for col in df.columns:
        html_parts.append(f'<th>{col}</th>')
    html_parts.append('</tr></thead>')
    
    # Data rows
    html_parts.append('<tbody>')
    for idx, (_, row) in enumerate(df.iterrows()):
        # Determine row class
        row_class = ""
        if idx in conflict_rows:
            row_class = "row-conflict"
        elif idx in new_rule_rows:
            row_class = "row-new-rule"
        
        html_parts.append(f'<tr class="{row_class}">')
        for val in row:
            # Handle NaN and None
            display_val = "" if pd.isna(val) else str(val)
            # Escape HTML
            display_val = display_val.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            html_parts.append(f'<td title="{display_val}">{display_val}</td>')
        html_parts.append('</tr>')
    html_parts.append('</tbody>')
    
    html_parts.append('</table>')
    html_parts.append('</div>')
    
    return ''.join(html_parts)


# =============================================================================
# MAIN TOOL FUNCTION
# =============================================================================

@tool
def display_csv(
    file_path: str,
    title: Optional[str] = None
) -> dict:
    """
    Display a CSV or Excel file in the UI artifact panel for human review.
    
    This tool renders the file as an interactive HTML table with:
    - Full data display (no truncation)
    - Row highlighting preserved from Excel files (red for conflicts, yellow for new rules)
    - Scrollable view for large files
    - Hover tooltips for truncated cell content
    
    IMPORTANT: This tool is for DISPLAY ONLY. Use this after processing to show
    results to the user for review. The UI will also provide a download button.
    
    Args:
        file_path: Full path to the CSV or Excel file to display
        title: Display title for the artifact panel (defaults to filename)
    
    Returns:
        Dictionary with display_type flag for UI artifact interception.
        The UI will render this as an interactive table with highlighting.
    """
    
    try:
        # Validate file exists
        if not os.path.exists(file_path):
            output = DisplayCsvOutput(
                display_type="dataframe",
                status='error',
                message=f"File not found: {file_path}",
                title=title or "Error",
                file_path=file_path,
                timestamp=datetime.now().isoformat(),
                error_type='FileNotFoundError',
                suggestion='Verify the file path is correct. Use find_file tool to locate the file.'
            )
            return output.model_dump()
        
        # Get absolute path
        abs_path = os.path.abspath(file_path)
        file_name = os.path.basename(abs_path)
        
        # Determine file type
        try:
            file_type = get_file_type(abs_path)
        except ValueError as e:
            output = DisplayCsvOutput(
                display_type="dataframe",
                status='error',
                message=str(e),
                title=title or "Error",
                file_path=abs_path,
                timestamp=datetime.now().isoformat(),
                error_type='UnsupportedFileTypeError',
                suggestion='Only CSV and Excel files (.csv, .xlsx, .xls, .xlsm) are supported.'
            )
            return output.model_dump()
        
        # Read file
        if file_type == 'csv':
            df = pd.read_csv(abs_path)
            highlight_data = {'conflict_rows': [], 'new_rule_rows': []}
        else:
            df = pd.read_excel(abs_path)
            # Extract highlighting from Excel
            highlight_data = extract_excel_highlighting(abs_path)
        
        # Get metadata
        file_size_mb = get_file_size_mb(abs_path)
        metadata = DisplayMetadata(
            total_rows=len(df),
            total_columns=len(df.columns),
            columns=df.columns.tolist(),
            file_size_mb=file_size_mb,
            file_type=file_type
        )
        
        # Create highlight info
        highlight_info = HighlightInfo(
            has_highlighting=bool(highlight_data['conflict_rows'] or highlight_data['new_rule_rows']),
            conflict_rows=highlight_data['conflict_rows'],
            new_rule_rows=highlight_data['new_rule_rows']
        )
        
        # Generate styled HTML
        display_title = title or file_name
        html_content = dataframe_to_styled_html(
            df=df,
            conflict_rows=highlight_data['conflict_rows'],
            new_rule_rows=highlight_data['new_rule_rows'],
            title=display_title
        )
        
        # Also keep CSV data for download
        csv_data = df.to_csv(index=False)
        
        # Create message
        message_parts = [f"Displaying {file_name} ({len(df)} rows, {len(df.columns)} columns)"]
        if highlight_info.has_highlighting:
            if highlight_info.conflict_rows:
                message_parts.append(f"🚨 {len(highlight_info.conflict_rows)} conflicts requiring review")
            if highlight_info.new_rule_rows:
                message_parts.append(f"📝 {len(highlight_info.new_rule_rows)} new rules for approval")
        
        output = DisplayCsvOutput(
            display_type="dataframe",
            status='success',
            message=". ".join(message_parts),
            title=display_title,
            file_path=abs_path,
            html_content=html_content,
            csv_data=csv_data,
            metadata=metadata,
            highlight_info=highlight_info,
            timestamp=datetime.now().isoformat()
        )
        
        return output.model_dump()
        
    except pd.errors.EmptyDataError:
        output = DisplayCsvOutput(
            display_type="dataframe",
            status='error',
            message='File is empty',
            title=title or "Error",
            file_path=file_path,
            timestamp=datetime.now().isoformat(),
            error_type='EmptyFileError',
            suggestion='The file contains no data.'
        )
        return output.model_dump()
        
    except Exception as e:
        output = DisplayCsvOutput(
            display_type="dataframe",
            status='error',
            message=f'Error displaying file: {str(e)}',
            title=title or "Error",
            file_path=file_path,
            timestamp=datetime.now().isoformat(),
            error_type=type(e).__name__,
            suggestion='Check that the file is a valid CSV or Excel file.'
        )
        return output.model_dump()


# =============================================================================
# END OF TOOL
# =============================================================================


# =============================================================================
# USAGE NOTES
# =============================================================================

"""
HOW THIS TOOL INTEGRATES WITH THE UI:

1. The tool returns a dict with "display_type": "dataframe"
2. The UI intercepts this in format_tool_result()
3. The html_content is rendered in the artifact panel
4. The csv_data is used for the download button
5. highlight_info tells the UI about conflict/new rule rows

AGENT USAGE:
- Call this tool AFTER processing to show results to user
- Use for: master contracts, delta files, merged contracts
- Always call for files that need human review

EXAMPLE:
    # After merge_and_highlight completes
    result = display_csv.invoke({
        "file_path": "/output/master_contract_merged_20250524.xlsx",
        "title": "Merged Contract - Review Required"
    })
"""


# =============================================================================
# FOR STANDALONE TESTING
# =============================================================================

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python display_csv_tool.py <file_path> [title]")
        sys.exit(1)
    
    file_path = sys.argv[1]
    title = sys.argv[2] if len(sys.argv) > 2 else None
    
    result = display_csv.invoke({
        "file_path": file_path,
        "title": title
    })
    
    print(f"\nStatus: {result['status']}")
    print(f"Message: {result['message']}")
    
    if result['status'] == 'success':
        print(f"\nFile: {result['file_path']}")
        print(f"Rows: {result['metadata']['total_rows']}")
        print(f"Columns: {result['metadata']['columns']}")
        
        if result['highlight_info']['has_highlighting']:
            print(f"\nHighlighting detected:")
            print(f"  Conflicts: {len(result['highlight_info']['conflict_rows'])} rows")
            print(f"  New Rules: {len(result['highlight_info']['new_rule_rows'])} rows")
        
        # Save HTML for preview
        with open("test_display.html", "w") as f:
            f.write(result['html_content'])
        print(f"\nHTML saved to: test_display.html")
    else:
        print(f"\nError: {result['error_type']}")
        print(f"Suggestion: {result['suggestion']}")