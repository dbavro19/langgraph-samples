"""
MERGE AND HIGHLIGHT - FUNCTION 3
Merge new rules with master contract and highlight for human review
"""

import pandas as pd
import os
from datetime import datetime
from typing import Tuple


# =============================================================================
# CONFIGURATION
# =============================================================================

# Hardcoded file paths for standalone execution
MASTER_CONTRACT_PATH = 'test_data_contracts.csv'
NEW_RULES_PATH = 'output_comparison/new_rules.csv'
OUTPUT_DIR = './output_final'


# =============================================================================
# FILE LOADING
# =============================================================================

def load_master_contract(csv_path: str) -> pd.DataFrame:
    """
    Load the master contract CSV
    
    Parameters:
        csv_path: Path to master contract CSV
        
    Returns:
        DataFrame with master contract rules
    """
    print(f"\n{'='*60}")
    print("STEP 1: Loading master contract")
    print(f"{'='*60}")
    
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"Master contract not found: {csv_path}")
    
    df = pd.read_csv(csv_path)
    print(f"✓ Loaded {len(df)} rules from master contract")
    
    return df


def load_new_rules(csv_path: str) -> pd.DataFrame:
    """
    Load the new rules CSV from Function 2
    
    Parameters:
        csv_path: Path to new_rules.csv
        
    Returns:
        DataFrame with new rules to add, or None if file doesn't exist
    """
    print(f"\n{'='*60}")
    print("STEP 2: Loading new rules (delta)")
    print(f"{'='*60}")
    
    if not os.path.exists(csv_path):
        print(f"⚠ Warning: New rules file not found: {csv_path}")
        print("  This likely means Function 2 hasn't been run yet.")
        return None
    
    df = pd.read_csv(csv_path)
    
    if len(df) == 0:
        print("✓ New rules file is empty - no changes needed!")
        return df
    else:
        print(f"✓ Loaded {len(df)} new rules to add")
        return df


# =============================================================================
# MERGE LOGIC
# =============================================================================

def merge_contracts(master_df: pd.DataFrame, new_rules_df: pd.DataFrame) -> pd.DataFrame:
    """
    Merge master contract with new rules
    
    Parameters:
        master_df: Master contract DataFrame
        new_rules_df: New rules DataFrame from Function 2
        
    Returns:
        Merged DataFrame with is_new_rule column
    """
    print(f"\n{'='*60}")
    print("STEP 3: Merging contracts")
    print(f"{'='*60}")
    
    # Prepare master contract - add is_new_rule column
    master_prepared = master_df.copy()
    master_prepared['is_new_rule'] = False
    
    # Prepare new rules - select only columns that exist in master
    # From Function 2, new_rules has: business_rule, business_term, rule_type, related_master_rule, confidence, reasoning
    # We only need: business_rule, business_term (and optionally Description if it exists)
    
    new_rules_prepared = pd.DataFrame()
    new_rules_prepared['Business Rule'] = new_rules_df['business_rule']
    new_rules_prepared['Business Term'] = new_rules_df['business_term']
    
    # Add Description column if it exists in master
    if 'Description' in master_prepared.columns:
        # Map rule_type to description
        def get_description(rule_type):
            descriptions = {
                'new_stricter': 'New requirement - stricter than existing',
                'new_different': 'New requirement - different constraint type',
                'new_business_term': 'New business term introduced',
                'conflict': 'New requirement - needs review (conflict)'
            }
            return descriptions.get(rule_type, 'New requirement')
        
        new_rules_prepared['Description'] = new_rules_df['rule_type'].apply(get_description)
    
    new_rules_prepared['is_new_rule'] = True
    
    # Merge
    merged_df = pd.concat([master_prepared, new_rules_prepared], ignore_index=True)
    
    # Sort by Business Term, then by is_new_rule (new rules after existing)
    merged_df = merged_df.sort_values(
        ['Business Term', 'is_new_rule'],
        ascending=[True, False]
    ).reset_index(drop=True)
    
    print(f"✓ Merged successfully")
    print(f"  - Existing rules: {len(master_prepared)}")
    print(f"  - New rules: {len(new_rules_prepared)}")
    print(f"  - Total rules: {len(merged_df)}")
    
    return merged_df


# =============================================================================
# OUTPUT GENERATION
# =============================================================================

def save_merged_contract(merged_df: pd.DataFrame, output_dir: str) -> Tuple[str, str]:
    """
    Save merged contract as CSV and Excel with highlighting
    
    Parameters:
        merged_df: Merged DataFrame with is_new_rule column
        output_dir: Output directory path
        
    Returns:
        Tuple of (csv_path, excel_path)
    """
    print(f"\n{'='*60}")
    print("STEP 4: Saving merged contract")
    print(f"{'='*60}")
    
    # Create output directory
    os.makedirs(output_dir, exist_ok=True)
    
    # Generate filename with date suffix
    date_suffix = datetime.now().strftime('%Y%m%d')
    base_filename = f'master_contract_with_proposed_changes_{date_suffix}'
    
    # Save CSV
    csv_path = os.path.join(output_dir, f'{base_filename}.csv')
    merged_df.to_csv(csv_path, index=False)
    print(f"✓ Saved CSV: {csv_path}")
    
    # Try to save Excel with highlighting
    excel_path = None
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font
        
        excel_path = os.path.join(output_dir, f'{base_filename}.xlsx')
        
        # Save to Excel first
        merged_df.to_excel(excel_path, index=False, engine='openpyxl')
        
        # Load workbook and apply highlighting
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Yellow fill for new rules
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        bold_font = Font(bold=True)
        
        # Find is_new_rule column
        is_new_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'is_new_rule':
                is_new_col = idx
                break
        
        # Apply yellow highlighting to new rules
        if is_new_col:
            new_rules_count = 0
            for row in range(2, ws.max_row + 1):
                is_new = ws.cell(row, is_new_col).value
                if is_new == True:
                    # Highlight entire row in yellow
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row, col).fill = yellow_fill
                    new_rules_count += 1
            
            print(f"✓ Applied yellow highlighting to {new_rules_count} new rules")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(excel_path)
        print(f"✓ Saved Excel with highlighting: {excel_path}")
        
    except ImportError:
        print(f"  ℹ openpyxl not available - skipping Excel export")
        print(f"    Install with: pip install openpyxl")
    
    return csv_path, excel_path


# =============================================================================
# MAIN WORKFLOW
# =============================================================================

def main():
    """Execute the merge and highlight workflow"""
    
    print("\n" + "="*60)
    print("MERGE AND HIGHLIGHT - FUNCTION 3")
    print("="*60)
    
    print("\nConfiguration:")
    print(f"  Master contract: {MASTER_CONTRACT_PATH}")
    print(f"  New rules: {NEW_RULES_PATH}")
    print(f"  Output directory: {OUTPUT_DIR}")
    
    # STEP 1: Load master contract
    try:
        master_df = load_master_contract(MASTER_CONTRACT_PATH)
    except FileNotFoundError as e:
        print(f"\n✗ ERROR: {e}")
        print("  Please ensure the master contract file exists.")
        return
    
    # STEP 2: Load new rules
    new_rules_df = load_new_rules(NEW_RULES_PATH)
    
    if new_rules_df is None:
        print(f"\n{'='*60}")
        print("TERMINATING - No new rules file found")
        print(f"{'='*60}")
        print("\n⚠ Please run Function 2 (contract_comparison.py) first to generate new_rules.csv")
        print("="*60 + "\n")
        return
    
    # Check if new rules file is empty
    if len(new_rules_df) == 0:
        print(f"\n{'='*60}")
        print("TERMINATING - No changes needed")
        print(f"{'='*60}")
        print("\n✓ The proposed contract is fully covered by the master contract.")
        print("✓ No new rules need to be added.")
        print("✓ No merge required.")
        print("="*60 + "\n")
        return
    
    # STEP 3: Merge contracts
    merged_df = merge_contracts(master_df, new_rules_df)
    
    # STEP 4: Save outputs
    csv_path, excel_path = save_merged_contract(merged_df, OUTPUT_DIR)
    
    # Final summary
    print(f"\n{'='*60}")
    print("MERGE COMPLETE - READY FOR HUMAN REVIEW")
    print(f"{'='*60}")
    
    new_rules_count = len(merged_df[merged_df['is_new_rule'] == True])
    existing_rules_count = len(merged_df[merged_df['is_new_rule'] == False])
    
    print(f"\nMerged Contract Summary:")
    print(f"  • Existing rules: {existing_rules_count}")
    print(f"  • New rules (highlighted): {new_rules_count}")
    print(f"  • Total rules: {len(merged_df)}")
    
    print(f"\n📄 Files Generated:")
    print(f"  • CSV: {csv_path}")
    if excel_path:
        print(f"  • Excel: {excel_path}")
        print(f"\n💡 Open the Excel file to review highlighted changes!")
        print(f"   Yellow rows = New rules requiring approval")
    
    print(f"\n→ Please review the highlighted rules and approve/reject changes.")
    print("="*60 + "\n")


if __name__ == "__main__":
    main()