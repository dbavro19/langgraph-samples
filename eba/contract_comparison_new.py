"""
CONTRACT COMPARISON - FUNCTION 2
Compare proposed contract against master contract to identify new rules needed

OVERVIEW:
When a consumer proposes a new data contract, we need to determine:
1. Which rules are already covered by the existing master contract
2. Which rules are genuinely new and need to be added (THE DELTA)
3. Avoid adding duplicate rules even within the proposed contract

This function performs DELTA DETECTION ONLY. It does NOT create the merged contract.
Function 3 will handle merging and highlighting for human review.
"""

import boto3
import json
import pandas as pd
import os
from typing import Dict, List, Any, Tuple


# =============================================================================
# CONFIGURATION
# =============================================================================

MAX_RETRIES = 3


# Import system prompts from separate module
from contract_comparison_system_prompt import (
    get_comparison_system_prompt,
    get_consolidation_system_prompt
)


# =============================================================================
# DATA LOADING
# =============================================================================

def load_master_contract(csv_path: str) -> pd.DataFrame:
    """
    Load the master (producer) contract CSV
    
    Parameters:
        csv_path: Path to master contract CSV
        
    Returns:
        DataFrame with columns: business_rule, Description, business_term
    """
    print(f"✓ Loading master contract from {csv_path}")
    df = pd.read_csv(csv_path)
    print(f"  Loaded {len(df)} master rules")
    return df


def load_proposed_contract(csv_path: str) -> pd.DataFrame:
    """
    Load the proposed (consumer) contract CSV
    
    Parameters:
        csv_path: Path to proposed contract CSV
        
    Returns:
        DataFrame with columns: business_rule, Description, business_term
    """
    print(f"✓ Loading proposed contract from {csv_path}")
    df = pd.read_csv(csv_path)
    print(f"  Loaded {len(df)} proposed rules")
    return df


# =============================================================================
# DATA VALIDATION AND CLEANING
# =============================================================================

def validate_and_clean_dataframe(df: pd.DataFrame, contract_type: str) -> pd.DataFrame:
    """
    Apply validation and quality checks to the dataframe
    
    Parameters:
        df: Raw input data
        contract_type: "master" or "proposed" (for logging)
        
    Returns:
        Cleaned and validated DataFrame
    """
    print(f"✓ Validating and cleaning {contract_type} contract")
    
    original_count = len(df)
    
    # Check required columns exist
    required_columns = ['business_rule', 'business_term']
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise ValueError(f"Missing required columns in {contract_type} contract: {missing_columns}")
    
    # Remove rows with null business_rule or business_term
    df = df.dropna(subset=['business_rule', 'business_term'])
    null_removed = original_count - len(df)
    if null_removed > 0:
        print(f"  Removed {null_removed} rows with null values")
    
    # Trim whitespace
    df['business_rule'] = df['business_rule'].str.strip()
    df['business_term'] = df['business_term'].str.strip()
    
    # Remove exact duplicates
    before_dedup = len(df)
    df = df.drop_duplicates(subset=['business_rule', 'business_term'])
    duplicates_removed = before_dedup - len(df)
    if duplicates_removed > 0:
        print(f"  Removed {duplicates_removed} exact duplicate rows")
    
    print(f"  Final {contract_type} dataset: {len(df)} rules")
    
    return df


# =============================================================================
# business_term EXTRACTION AND CATEGORIZATION
# =============================================================================

def categorize_business_terms(master_df: pd.DataFrame, proposed_df: pd.DataFrame) -> Dict[str, List[str]]:
    """
    Categorize business_terms into existing and new
    
    Parameters:
        master_df: Cleaned master contract
        proposed_df: Cleaned proposed contract
        
    Returns:
        Dictionary with categorized terms
    """
    print("✓ Categorizing business_terms")
    
    # Get unique business_terms from each contract
    master_terms = set(master_df['business_term'].unique())
    proposed_terms = set(proposed_df['business_term'].unique())
    
    # Find existing terms (intersection)
    existing_terms = list(master_terms.intersection(proposed_terms))
    existing_terms.sort()
    
    # Find new terms (only in proposed)
    new_terms = list(proposed_terms - master_terms)
    new_terms.sort()
    
    print(f"  Existing terms (in both contracts): {len(existing_terms)}")
    print(f"  New terms (only in proposed): {len(new_terms)}")
    
    return {
        "existing_terms": existing_terms,
        "new_terms": new_terms
    }


def get_rules_for_business_term(df: pd.DataFrame, business_term: str) -> pd.DataFrame:
    """
    Filter dataframe to only rules for a specific business_term
    
    Parameters:
        df: Contract data
        business_term: The business_term to filter for
        
    Returns:
        DataFrame subset containing only rows for this business_term
    """
    return df[df['business_term'] == business_term].copy()


# =============================================================================
# LLM INVOCATION
# =============================================================================

def invoke_bedrock_model(system_prompt: str, user_prompt: str) -> str:
    """
    Call AWS Bedrock API to invoke Claude
    
    Parameters:
        system_prompt: System instructions
        user_prompt: User message
        
    Returns:
        Raw response text from model
    """
    # Initialize Bedrock client
    bedrock = boto3.client(
        service_name='bedrock-runtime',
        region_name='us-east-1'
    )
    
    # Build request
    body = {
        "modelId": "us.anthropic.claude-haiku-4-5-20251001-v1:0",
        "inferenceConfig": {
            "maxTokens": 4096,
            "temperature": 0
        },
        "system": [{"text": system_prompt}],
        "messages": [{
            'role': 'user',
            'content': [{'text': user_prompt}]
        }]
    }
    
    # Call API
    response = bedrock.converse(**body)
    
    # Extract text
    result = response['output']['message']['content'][0]['text']
    
    return result


# =============================================================================
# LLM COMPARISON - EXISTING business_termS
# =============================================================================

def compare_rules_for_existing_term(master_rules_df: pd.DataFrame, 
                                   proposed_rules_df: pd.DataFrame, 
                                   business_term: str) -> Dict[str, Any]:
    """
    Compare proposed rules against master rules for an existing business_term
    Uses LLM to determine which proposed rules are new vs already covered
    
    Parameters:
        master_rules_df: All master rules for this business_term
        proposed_rules_df: All proposed rules for this business_term
        business_term: The business_term being compared
        
    Returns:
        Dictionary with structured JSON comparison analysis
    """
    # Extract rule lists
    master_rules = master_rules_df['business_rule'].tolist()
    proposed_rules = proposed_rules_df['business_rule'].tolist()
    
    # Build user prompt
    user_prompt = f"""business_term: {business_term}

Master Contract Rules:
{json.dumps(master_rules, indent=2)}

Proposed Contract Rules:
{json.dumps(proposed_rules, indent=2)}

Compare each proposed rule against the master rules. For each proposed rule, determine:
1. Is it a duplicate of another proposed rule?
2. Is it already covered by a master rule (same or stricter)?
3. Is it a new requirement that needs to be added?
4. Does it conflict with master rules?

Return your analysis as valid JSON only (no markdown, no code blocks).

Required JSON structure:
{{
    "business_term": "{business_term}",
    "master_rules_count": {len(master_rules)},
    "proposed_rules_count": {len(proposed_rules)},
    "new_rules_required": <boolean>,
    "analysis": [
        {{
            "proposed_rule": "<exact proposed rule text>",
            "proposed_rule_index": <index>,
            "status": "covered|new_stricter|new_different|conflict|duplicate_within_proposed",
            "matched_master_rule": "<master rule text or null>",
            "matched_master_index": <index or null>,
            "canonical_rule": "<rule to add if new, or master rule if covered>",
            "confidence": <0.0 to 1.0>,
            "reasoning": "<clear explanation>"
        }}
    ],
    "notes": "<any observations>"
}}"""
    
    # Get system prompt
    system_prompt = get_comparison_system_prompt()
    
    # Attempt with retries
    for attempt in range(MAX_RETRIES):
        try:
            print(f"    → Calling LLM for comparison (attempt {attempt + 1}/{MAX_RETRIES})...")
            
            # Call LLM
            llm_response = invoke_bedrock_model(system_prompt, user_prompt)
            
            # Clean and parse response
            llm_response = llm_response.strip()
            if llm_response.startswith("```"):
                lines = llm_response.split('\n')
                llm_response = '\n'.join(lines[1:-1])
            
            result = json.loads(llm_response)
            
            # Validate structure
            validate_comparison_response_structure(result, business_term)
            
            print(f"    ✓ Successfully parsed comparison response")
            return result
            
        except (json.JSONDecodeError, ValueError, KeyError) as e:
            print(f"    ✗ Attempt {attempt + 1} failed: {str(e)}")
            
            if attempt < MAX_RETRIES - 1:
                user_prompt += "\n\nIMPORTANT: Return ONLY valid JSON. No markdown code blocks, no explanations outside JSON."
            else:
                print(f"    ✗ ERROR: Failed to parse comparison response after {MAX_RETRIES} attempts")
                return {
                    "business_term": business_term,
                    "error": f"Failed to parse LLM response: {str(e)}",
                    "master_rules_count": len(master_rules),
                    "proposed_rules_count": len(proposed_rules),
                    "new_rules_required": False,
                    "analysis": []
                }


def validate_comparison_response_structure(response: Dict, business_term: str):
    """
    Validate that comparison LLM response has expected structure
    
    Parameters:
        response: Parsed JSON from LLM
        business_term: Expected business_term
        
    Raises:
        ValueError if validation fails
    """
    # Check required fields
    required_fields = ['business_term', 'analysis', 'new_rules_required']
    for field in required_fields:
        if field not in response:
            raise ValueError(f"Missing '{field}' field")
    
    if not isinstance(response['analysis'], list):
        raise ValueError("'analysis' must be a list")
    
    # Validate each analysis entry
    for i, entry in enumerate(response['analysis']):
        required_entry_fields = ['proposed_rule', 'status', 'canonical_rule', 'confidence']
        for field in required_entry_fields:
            if field not in entry:
                raise ValueError(f"Analysis entry {i} missing field: {field}")
        
        # Check status value
        valid_statuses = ['covered', 'new_stricter', 'new_different', 'conflict', 'duplicate_within_proposed']
        if entry['status'] not in valid_statuses:
            raise ValueError(f"Analysis entry {i} has invalid status: {entry['status']}")
        
        # Check confidence range
        if not (0.0 <= entry['confidence'] <= 1.0):
            raise ValueError(f"Analysis entry {i} has invalid confidence: {entry['confidence']}")


# =============================================================================
# SELF-CONSOLIDATION - NEW business_termS
# =============================================================================

def consolidate_rules_for_new_term(proposed_rules_df: pd.DataFrame, business_term: str) -> Dict[str, Any]:
    """
    Consolidate proposed rules for a NEW business_term (not in master)
    Uses Function 1 logic to identify duplicates within proposed rules
    
    Parameters:
        proposed_rules_df: All proposed rules for this new term
        business_term: The business_term
        
    Returns:
        Dictionary with structured JSON consolidation analysis
    """
    # Extract list of proposed rules
    rules_list = proposed_rules_df['business_rule'].tolist()
    
    # If only 1 rule, no LLM call needed
    if len(rules_list) == 1:
        return {
            "business_term": business_term,
            "total_rules_analyzed": 1,
            "unique_rules_identified": 1,
            "duplicates_found": False,
            "analysis": [
                {
                    "original_rule": rules_list[0],
                    "original_rule_index": 0,
                    "status": "unique",
                    "canonical_rule": rules_list[0],
                    "confidence": 1.0,
                    "consolidated_with": [],
                    "duplicate_of_index": None,
                    "reasoning": "Only one rule for this new business_term. No consolidation needed."
                }
            ],
            "notes": "Single rule for new business_term - no LLM processing required"
        }
    
    # Multiple rules - use LLM for consolidation
    user_prompt = f"""business_term: {business_term}

Rules to analyze:
{json.dumps(rules_list, indent=2)}

Analyze these rules and identify semantic duplicates. Return your analysis as valid JSON only (no markdown, no code blocks).

Required JSON structure:
{{
    "business_term": "{business_term}",
    "total_rules_analyzed": {len(rules_list)},
    "unique_rules_identified": <number>,
    "duplicates_found": <boolean>,
    "analysis": [
        {{
            "original_rule": "<exact rule text>",
            "original_rule_index": <index>,
            "status": "unique" or "duplicate",
            "canonical_rule": "<best version>",
            "confidence": <0.0 to 1.0>,
            "consolidated_with": [<array of duplicate rule texts>],
            "duplicate_of_index": <index or null>,
            "reasoning": "<clear explanation>"
        }}
    ],
    "notes": "<any observations>"
}}"""
    
    # Get system prompt
    system_prompt = get_consolidation_system_prompt()
    
    # Attempt with retries
    for attempt in range(MAX_RETRIES):
        try:
            print(f"    → Calling LLM for consolidation (attempt {attempt + 1}/{MAX_RETRIES})...")
            
            # Call LLM
            llm_response = invoke_bedrock_model(system_prompt, user_prompt)
            
            # Clean and parse response
            llm_response = llm_response.strip()
            if llm_response.startswith("```"):
                lines = llm_response.split('\n')
                llm_response = '\n'.join(lines[1:-1])
            
            result = json.loads(llm_response)
            
            # Validate structure (reuse validation from consolidate.py)
            validate_consolidation_response_structure(result, business_term)
            
            print(f"    ✓ Successfully parsed consolidation response")
            return result
            
        except (json.JSONDecodeError, ValueError, KeyError) as e:
            print(f"    ✗ Attempt {attempt + 1} failed: {str(e)}")
            
            if attempt < MAX_RETRIES - 1:
                user_prompt += "\n\nIMPORTANT: Return ONLY valid JSON. No markdown code blocks, no explanations outside JSON."
            else:
                print(f"    ✗ ERROR: Failed to parse consolidation response after {MAX_RETRIES} attempts")
                return {
                    "business_term": business_term,
                    "error": f"Failed to parse LLM response: {str(e)}",
                    "total_rules_analyzed": len(rules_list),
                    "unique_rules_identified": 0,
                    "duplicates_found": False,
                    "analysis": []
                }


def validate_consolidation_response_structure(response: Dict, business_term: str):
    """
    Validate that consolidation LLM response has expected structure
    
    Parameters:
        response: Parsed JSON from LLM
        business_term: Expected business_term
        
    Raises:
        ValueError if validation fails
    """
    # Check required fields
    if 'business_term' not in response:
        raise ValueError("Missing 'business_term' field")
    
    if 'analysis' not in response:
        raise ValueError("Missing 'analysis' field")
    
    if not isinstance(response['analysis'], list):
        raise ValueError("'analysis' must be a list")
    
    # Validate each analysis entry
    for i, entry in enumerate(response['analysis']):
        required_fields = ['original_rule', 'status', 'canonical_rule', 'confidence']
        for field in required_fields:
            if field not in entry:
                raise ValueError(f"Analysis entry {i} missing field: {field}")
        
        # Check status value
        if entry['status'] not in ['unique', 'duplicate']:
            raise ValueError(f"Analysis entry {i} has invalid status: {entry['status']}")
        
        # Check confidence range
        if not (0.0 <= entry['confidence'] <= 1.0):
            raise ValueError(f"Analysis entry {i} has invalid confidence: {entry['confidence']}")


# =============================================================================
# ORCHESTRATION
# =============================================================================

def process_all_business_terms(master_df: pd.DataFrame, 
                              proposed_df: pd.DataFrame,
                              categorized_terms: Dict[str, List[str]]) -> List[Dict[str, Any]]:
    """
    Process all business_terms - both existing (comparison) and new (consolidation)
    
    Parameters:
        master_df: Cleaned master contract
        proposed_df: Cleaned proposed contract
        categorized_terms: Dictionary with existing_terms and new_terms lists
        
    Returns:
        List of analysis results for all terms
    """
    print(f"\n{'='*60}")
    print("STEP 3: Processing business_terms")
    print(f"{'='*60}")
    
    all_results = []
    
    # Process existing terms (comparison against master)
    existing_terms = categorized_terms["existing_terms"]
    if existing_terms:
        print(f"\nProcessing {len(existing_terms)} existing terms (comparison against master):")
        for i, business_term in enumerate(existing_terms, 1):
            master_rules_df = get_rules_for_business_term(master_df, business_term)
            proposed_rules_df = get_rules_for_business_term(proposed_df, business_term)
            
            print(f"\n[{i}/{len(existing_terms)}] {business_term}")
            print(f"  Master rules: {len(master_rules_df)}, Proposed rules: {len(proposed_rules_df)}")
            
            # Compare proposed against master
            result = compare_rules_for_existing_term(master_rules_df, proposed_rules_df, business_term)
            result['term_type'] = 'existing'
            all_results.append(result)
    
    # Process new terms (consolidation within proposed)
    new_terms = categorized_terms["new_terms"]
    if new_terms:
        print(f"\n\nProcessing {len(new_terms)} new terms (consolidation within proposed):")
        for i, business_term in enumerate(new_terms, 1):
            proposed_rules_df = get_rules_for_business_term(proposed_df, business_term)
            
            print(f"\n[{i}/{len(new_terms)}] {business_term}")
            print(f"  Proposed rules: {len(proposed_rules_df)}")
            
            # Consolidate within proposed rules
            result = consolidate_rules_for_new_term(proposed_rules_df, business_term)
            result['term_type'] = 'new'
            all_results.append(result)
    
    return all_results


# =============================================================================
# OUTPUT GENERATION
# =============================================================================

def create_delta_outputs(all_results: List[Dict[str, Any]]) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, Any]]:
    """
    Create delta outputs from analysis results
    
    Parameters:
        all_results: Analysis results from all business_terms
        
    Returns:
        Tuple of (new_rules_df, comparison_audit_df, summary_dict)
    """
    print(f"\n{'='*60}")
    print("STEP 4: Creating delta outputs")
    print(f"{'='*60}")
    
    new_rules = []
    audit_records = []
    
    # Process each result
    for result in all_results:
        business_term = result['business_term']
        term_type = result.get('term_type', 'unknown')
        
        # Skip if error occurred
        if 'error' in result:
            print(f"  ⚠ Warning: Skipping {business_term} due to error: {result['error']}")
            continue
        
        # Track canonical rules we've seen for this term
        seen_canonical = set()
        
        # Process each rule analysis
        for rule_analysis in result['analysis']:
            # Create audit record
            audit_record = {
                'business_term': business_term,
                'term_type': term_type,
                'original_rule': rule_analysis.get('original_rule', rule_analysis.get('proposed_rule', '')),
                'status': rule_analysis['status'],
                'canonical_rule': rule_analysis['canonical_rule'],
                'confidence': rule_analysis['confidence'],
                'needs_review': rule_analysis['confidence'] < 0.9,
                'reasoning': rule_analysis['reasoning']
            }
            
            # Add comparison-specific fields for existing terms
            if term_type == 'existing':
                audit_record['matched_master_rule'] = rule_analysis.get('matched_master_rule')
                audit_record['matched_master_index'] = rule_analysis.get('matched_master_index')
            
            audit_records.append(audit_record)
            
            # Add to new rules if needed
            canonical_rule = rule_analysis['canonical_rule']
            status = rule_analysis['status']
            
            # Rules that need to be added to master contract
            if status in ['new_stricter', 'new_different'] or (term_type == 'new' and status == 'unique'):
                if canonical_rule not in seen_canonical:
                    seen_canonical.add(canonical_rule)
                    
                    # Count original rules mapping to this canonical
                    original_rule_count = sum(
                        1 for r in result['analysis'] 
                        if r['canonical_rule'] == canonical_rule
                    )
                    
                    new_rules.append({
                        'business_rule': canonical_rule,
                        'business_term': business_term,
                        'term_type': term_type,
                        'rule_source': 'comparison_analysis',
                        'original_rule_count': original_rule_count,
                        'confidence': rule_analysis['confidence'],
                        'needs_review': rule_analysis['confidence'] < 0.9
                    })
    
    # Convert to DataFrames
    new_rules_df = pd.DataFrame(new_rules)
    audit_df = pd.DataFrame(audit_records)
    
    # Sort for consistency
    if not new_rules_df.empty:
        new_rules_df = new_rules_df.sort_values(['business_term', 'business_rule']).reset_index(drop=True)
    
    if not audit_df.empty:
        audit_df = audit_df.sort_values(
            ['business_term', 'needs_review', 'original_rule'],
            ascending=[True, False, True]
        ).reset_index(drop=True)
    
    # Create summary
    summary = {
        'total_proposed_rules': len(audit_df),
        'new_rules_identified': len(new_rules_df),
        'rules_already_covered': len(audit_df[audit_df['status'] == 'covered']) if not audit_df.empty else 0,
        'rules_needing_review': len(audit_df[audit_df['needs_review'] == True]) if not audit_df.empty else 0,
        'business_terms_analyzed': len(audit_df['business_term'].unique()) if not audit_df.empty else 0,
        'existing_terms_compared': len([r for r in all_results if r.get('term_type') == 'existing']),
        'new_terms_consolidated': len([r for r in all_results if r.get('term_type') == 'new'])
    }
    
    print(f"✓ Created new rules list with {len(new_rules_df)} rules to add")
    print(f"✓ Created comparison audit with {len(audit_df)} total rules analyzed")
    
    return new_rules_df, audit_df, summary


def save_outputs(new_rules_df: pd.DataFrame, 
                audit_df: pd.DataFrame, 
                summary: Dict[str, Any], 
                output_dir: str = './output') -> Dict[str, Any]:
    """
    Save contract comparison outputs
    
    Parameters:
        new_rules_df: Rules that need to be added (the delta)
        audit_df: Detailed comparison audit trail
        summary: Summary statistics
        output_dir: Directory to save outputs
        
    Returns:
        Summary dictionary
    """
    print(f"\n{'='*60}")
    print("STEP 5: Saving outputs")
    print(f"{'='*60}")
    
    # Create output directory
    os.makedirs(output_dir, exist_ok=True)
    
    # Save new rules (the delta)
    new_rules_path = os.path.join(output_dir, 'new_rules_delta.csv')
    new_rules_df.to_csv(new_rules_path, index=False)
    print(f"✓ Saved new rules delta: {new_rules_path}")
    
    # Save comparison audit
    audit_path = os.path.join(output_dir, 'comparison_audit.csv')
    audit_df.to_csv(audit_path, index=False)
    print(f"✓ Saved comparison audit: {audit_path}")
    
    # Try to save Excel with formatting
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font
        
        excel_path = os.path.join(output_dir, 'comparison_audit_formatted.xlsx')
        audit_df.to_excel(excel_path, index=False, engine='openpyxl')
        
        # Apply formatting
        wb = load_workbook(excel_path)
        ws = wb.active
        
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        # Find needs_review column
        needs_review_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'needs_review':
                needs_review_col = idx
                break
        
        # Highlight rows where needs_review = TRUE
        if needs_review_col:
            for row in range(2, ws.max_row + 1):
                if ws.cell(row, needs_review_col).value == True:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row, col).fill = yellow_fill
        
        wb.save(excel_path)
        print(f"✓ Saved formatted audit trail: {excel_path}")
        
    except ImportError:
        print(f"  ℹ openpyxl not available - skipping Excel export")
    
    # Save summary
    summary_path = os.path.join(output_dir, 'comparison_summary.json')
    with open(summary_path, 'w') as f:
        json.dump(summary, f, indent=2)
    print(f"✓ Saved summary: {summary_path}")
    
    return summary


# =============================================================================
# MAIN WORKFLOW
# =============================================================================

def main(master_contract_path: str, proposed_contract_path: str, output_dir: str = './output'):
    """Execute the complete contract comparison workflow"""
    
    print("\n" + "="*60)
    print("CONTRACT COMPARISON - DELTA DETECTION")
    print("="*60)
    
    # STEP 1: Load data
    print(f"\n{'='*60}")
    print("STEP 1: Loading contract data")
    print(f"{'='*60}")
    
    master_df = load_master_contract(master_contract_path)
    proposed_df = load_proposed_contract(proposed_contract_path)
    
    # STEP 2: Validate and clean
    print(f"\n{'='*60}")
    print("STEP 2: Validating and cleaning data")
    print(f"{'='*60}")
    
    master_df_clean = validate_and_clean_dataframe(master_df, "master")
    proposed_df_clean = validate_and_clean_dataframe(proposed_df, "proposed")
    
    # Categorize business_terms
    categorized_terms = categorize_business_terms(master_df_clean, proposed_df_clean)
    
    print(f"\nProcessing Summary:")
    print(f"  • Existing terms (in both contracts): {len(categorized_terms['existing_terms'])}")
    print(f"  • New terms (only in proposed): {len(categorized_terms['new_terms'])}")
    print(f"  • Total proposed rules: {len(proposed_df_clean)}")
    print(f"  • Total master rules: {len(master_df_clean)}")
    
    # STEP 3: Process all business_terms
    all_results = process_all_business_terms(master_df_clean, proposed_df_clean, categorized_terms)
    
    # STEP 4: Create delta outputs
    new_rules_df, audit_df, summary = create_delta_outputs(all_results)
    
    # STEP 5: Save outputs
    summary = save_outputs(new_rules_df, audit_df, summary, output_dir)
    
    # Final summary
    print(f"\n{'='*60}")
    print("COMPARISON COMPLETE")
    print(f"{'='*60}")
    print(f"\nResults:")
    print(f"  • Total proposed rules analyzed: {summary['total_proposed_rules']}")
    print(f"  • New rules identified (THE DELTA): {summary['new_rules_identified']}")
    print(f"  • Rules already covered by master: {summary['rules_already_covered']}")
    print(f"  • business_terms analyzed: {summary['business_terms_analyzed']}")
    
    if summary['rules_needing_review'] > 0:
        print(f"\n⚠️  {summary['rules_needing_review']} rules flagged for review (confidence < 0.9)")
        print(f"   Review in: {output_dir}/comparison_audit_formatted.xlsx (highlighted rows)")
    
    print(f"\n✓ All outputs saved to {output_dir}/")
    print("="*60 + "\n")
    
    return summary


if __name__ == "__main__":
    # Example usage
    main(
        master_contract_path='output/master_contract.csv',
        proposed_contract_path='test_compare_contract.csv',
        output_dir='./output'
    )